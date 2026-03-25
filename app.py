from tkinter import ON

from flask import Flask, render_template, request, redirect, url_for, send_file, flash
from flask_login import login_user, login_required, logout_user, current_user
from auth import login_manager, users, User

import mysql.connector
import pandas as pd
from datetime import datetime, date, timedelta
import re
import json
from collections import OrderedDict
from flask import request
from flask import send_file
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from urllib.parse import quote
import io
app = Flask(__name__)
app.secret_key = "secret123"

login_manager.init_app(app)
login_manager.login_view = "login"

PACKAGE_TUNA_MENU_NAMES = (
    "dada tuna goreng",
    "dada tuna bakar",
    "paket dada tuna goreng",
    "paket dada tuna bakar"
)
SPECIAL_TUNA_MENU_NAMES = PACKAGE_TUNA_MENU_NAMES + ("rahang tuna",)
FRIED_RICE_MENU_NAMES = (
    "nasi goreng kampung",
    "nasi goreng spesial",
    "nasi goreng sagela"
)
DASHBOARD_TABLE_CONFIGS = OrderedDict([
    ("fish_stock", {
        "label": "Total Ikan",
        "title": "Total Ikan Hari Ini",
        "description": "Tabel ini terpisah dari filter divisi dan dihitung berdasarkan tanggal yang dipilih.",
        "empty_message": "Belum ada total ikan pada tanggal ini.",
        "table_class": "w-full text-left",
        "excel_title": "TOTAL STOCK IKAN YANG DIBUTUHKAN HARI INI",
        "excel_table_name": "RekapIkan",
        "excel_style_name": "TableStyleMedium4",
        "compact_number_keys": {"no", "total"},
        "columns": [
            {"key": "no", "label": "No"},
            {"key": "fish_divisi", "label": "Divisi"},
            {"key": "fish_name", "label": "Jenis Ikan"},
            {"key": "fish_size", "label": "Ukuran"},
            {"key": "fish_weight_display", "label": "Berat"},
            {"key": "total", "label": "Jumlah"},
            {"key": "menu_names", "label": "Menu"}
        ]
    }),
    ("rice_need", {
        "label": "Kebutuhan Nasi",
        "title": "Rekap Kebutuhan Nasi Hari Ini",
        "description": "",
        "empty_message": "Belum ada kebutuhan nasi pada tanggal ini.",
        "table_class": "w-full text-left",
        "excel_title": "REKAP KEBUTUHAN NASI HARI INI",
        "excel_table_name": "RekapNasi",
        "excel_style_name": "TableStyleMedium2",
        "compact_number_keys": {"no", "jumlah"},
        "columns": [
            {"key": "no", "label": "No"},
            {"key": "keterangan", "label": "Keterangan"},
            {"key": "jumlah", "label": "Jumlah"}
        ]
    }),
    ("fried_rice", {
        "label": "Total Nasi Goreng",
        "title": "Total Sambal Nasi Goreng",
        "description": "",
        "empty_message": "Belum ada total nasi goreng pada tanggal ini.",
        "table_class": "w-full text-left",
        "excel_title": "TOTAL SAMBAL NASI GORENG",
        "excel_table_name": "RekapNasiGoreng",
        "excel_style_name": "TableStyleMedium3",
        "compact_number_keys": {"no", "jumlah"},
        "columns": [
            {"key": "no", "label": "No"},
            {"key": "menu", "label": "Menu"},
            {"key": "jumlah", "label": "Jumlah"}
        ]
    }),
    ("menu_total", {
        "label": "Total Jumlah Makanan Permenu",
        "title": "TOTAL MENU UNTUK KITCHEN",
        "description": "Menu digabung per nama dasar. PCS dan paket dijumlahkan menjadi satu total, dengan ringkasan detail order, penyajian saus sate, dan special request.",
        "empty_message": "Belum ada total menu kitchen pada tanggal ini.",
        "table_class": "w-full text-left",
        "excel_title": "TOTAL MENU UNTUK KITCHEN",
        "excel_table_name": "RekapMenuGabungan",
        "excel_style_name": "TableStyleMedium6",
        "compact_number_keys": {"no", "jumlah"},
        "columns": [
            {"key": "no", "label": "No"},
            {"key": "divisi", "label": "Divisi"},
            {"key": "menu", "label": "Nama Menu"},
            {"key": "jumlah", "label": "Jumlah"},
            {"key": "special_req", "label": "Special Req"},
            {"key": "penyajian_saus", "label": "Keterangan Saus"},
            {"key": "detail", "label": "Keterangan"}
        ]
    }),
    ("menu_total_detail", {
        "label": "Total Menu Detail",
        "title": "TOTAL MENU UNTUK PLATING",
        "description": "Menu dipisah per tipe serving seperti PCS, Paket, atau Porsi, lengkap dengan detail order, penyajian saus sate, dan special request.",
        "empty_message": "Belum ada detail menu pada tanggal ini.",
        "table_class": "w-full text-left",
        "excel_title": "TOTAL MENU UNTUK PLATING",
        "excel_table_name": "RekapMenuDetail",
        "excel_style_name": "TableStyleMedium7",
        "compact_number_keys": {"no", "jumlah"},
        "columns": [
            {"key": "no", "label": "No"},
            {"key": "divisi", "label": "Divisi"},
            {"key": "menu", "label": "Menu"},
            {"key": "jumlah", "label": "Jumlah"},
            {"key": "special_req", "label": "Special Req"},
            {"key": "penyajian_saus", "label": "Penyajian Saus"},
            {"key": "detail", "label": "Keterangan"},
        ]
    })
])
EXPORT_TABLE_KEYS = set(DASHBOARD_TABLE_CONFIGS.keys())
RESTAURANT_PROFILE = {
    "name": "Manna Bakery and Cafe",
    "tagline": "Tempat nyaman untuk ngopi, makan santai, acara keluarga, dan reservasi rombongan.",
    "description": "Laman ini dibuat sebagai tampilan publik agar tamu bisa booking meja, melihat alamat, dan langsung chat WhatsApp Manna Bakery and Cafe.",
    "address": "Jl. Taman Bunga, Moodu, Kec. Kota Tim., Kota Gorontalo, Gorontalo, Indonesia",
    "maps_url": "https://www.google.com/maps/place/Manna+Bakery+%26+Cafe/@0.5422788,123.0706622,18z/data=!4m15!1m8!3m7!1s0x32792b2581d491e9:0xe903578a06130f2f!2sJl.+Taman+Bunga,+Kec.+Kota+Tim.,+Kota+Gorontalo,+Gorontalo+96135!3b1!8m2!3d0.5433483!4d123.07236!16s%2Fg%2F11cn8tf56c!3m5!1s0x32792ba4ec34408f:0x5e8e69efd821cdbf!8m2!3d0.5415071!4d123.0705396!16s%2Fg%2F11j3q8ccyk?entry=ttu&g_ep=EgoyMDI2MDMyMi4wIKXMDSoASAFQAw%3D%3D",
    "whatsapp_number": "628113112919",
    "whatsapp_display": "08113112919",
    "hours": [
        {"label": "Senin - Jumat", "value": "10.00 - 22.00"},
        {"label": "Sabtu - Minggu", "value": "09.00 - 23.00"},
        {"label": "Reservasi Grup", "value": "Sebaiknya booking lebih awal"}
    ],
    "features": [
        {
            "title": "Booking Mudah",
            "description": "Isi form singkat untuk kirim permintaan reservasi langsung ke sistem restoran."
        },
        {
            "title": "Chat WhatsApp",
            "description": "Tamu bisa langsung menghubungi restoran untuk konfirmasi cepat atau tanya ketersediaan."
        },
        {
            "title": "Info Lokasi",
            "description": "Alamat, jam operasional, dan akses ke Google Maps tersedia dalam satu halaman."
        }
    ],
    "highlights": [
        "Area makan nyaman untuk keluarga dan rombongan.",
        "Bisa dipakai untuk acara kecil, meeting, dan dinner santai.",
        "Admin tetap dapat masuk ke panel operasional restoran dari laman ini."
    ]
}


def build_option_group(key, label, choices, default=None, multiple=False):
    choice_rows = [{"value": value, "label": choice_label} for value, choice_label in choices]
    if multiple:
        if default is None:
            default_value = [choice_rows[0]["value"]] if choice_rows else []
        elif isinstance(default, (list, tuple, set)):
            default_value = [str(value).strip() for value in default if str(value).strip()]
        else:
            default_value = [str(default).strip()] if str(default).strip() else []
    else:
        default_value = default or (choice_rows[0]["value"] if choice_rows else "")

    return {
        "key": key,
        "label": label,
        "default": default_value,
        "multiple": multiple,
        "choices": choice_rows
    }


SERVING_TYPE_OPTIONS = build_option_group(
    "serving_type",
    "Pilih Paket atau PCS",
    [("paket", "Paket"), ("pcs", "PCS")],
    default="pcs"
)
CHICKEN_GEPREK_PART_OPTIONS = build_option_group(
    "chicken_part",
    "Bagian Ayam",
    [("paha_bawah", "Paha Bawah"), ("paha_atas", "Paha Atas")]
)
CHICKEN_STANDARD_PART_OPTIONS = build_option_group(
    "chicken_part",
    "Bagian Ayam",
    [("dada", "Dada"), ("paha", "Paha")]
)
GEPREK_RICA_OPTIONS = build_option_group(
    "rica_mode",
    "Sambal Rica",
    [("langsung", "Langsung"), ("pisah", "Pisah Rica")]
)
TUNA_COOKING_OPTIONS = build_option_group(
    "cook_style",
    "Cara Masak",
    [("bakar", "Bakar"), ("goreng", "Goreng")]
)
EGG_STYLE_OPTIONS = build_option_group(
    "egg_style",
    "Pilihan Telur",
    [("dadar", "Dadar"), ("rebus", "Rebus"), ("mata_sapi", "Mata Sapi")]
)
FRIED_RICE_EGG_OPTIONS = build_option_group(
    "egg_style",
    "Pilihan Telur",
    [
        ("ceplok_setengah_matang", "Ceplok Setengah Matang"),
        ("ceplok_matang", "Ceplok Matang"),
        ("rebus", "Rebus"),
        ("dadar", "Dadar")
    ]
)
BROTH_STYLE_OPTIONS = build_option_group(
    "broth_style",
    "Pilihan Kuah",
    [("pisah", "Pisah Kuah"), ("campur", "Campur Kuah")]
)
BANANA_TYPE_OPTIONS = build_option_group(
    "banana_type",
    "Jenis Pisang",
    [("raja", "Raja"), ("pagata", "Pagata")]
)
SERUT_STYLE_OPTIONS = build_option_group(
    "serut_style",
    "Pilihan Siraman",
    [("gulmer", "Gulmer"), ("sirup", "Sirup")]
)
BIG_FRIED_STYLE_OPTIONS = build_option_group(
    "fried_style",
    "Varian Ayam",
    [("ori", "Ori"), ("bubble", "Bubble")]
)
BIG_FRIED_SIZE_OPTIONS = build_option_group(
    "fried_size",
    "Ukuran",
    [("small", "Small"), ("large", "Large")]
)
BIG_FRIED_SEASONING_OPTIONS = build_option_group(
    "seasoning",
    "Bumbu",
    [
        ("balado", "Balado"),
        ("keju", "Keju"),
        ("bbq", "BBQ"),
        ("jagung_bakar", "Jagung Bakar"),
        ("sapi_panggang", "Sapi Panggang"),
        ("extra_hot", "Extra Hot")
    ],
    multiple=True
)
TOPPING_SAUCE_OPTIONS = build_option_group(
    "topping_sauce",
    "Topping Saus",
    [
        ("saus_keju", "Saus Keju"),
        ("saus_mayo", "Saus Mayo"),
        ("saus_bbq", "Saus BBQ"),
        ("saus_pedas_manis", "Saus Pedas Manis"),
        ("saus_rica_bawang", "Saus Rica Bawang")
    ],
    multiple=True
)
TEMPERATURE_OPTIONS = build_option_group(
    "temperature",
    "Suhu Minuman",
    [("ice", "Ice"), ("hangat", "Hangat"), ("hot", "Hot")]
)
FRUIT_JUICE_OPTIONS = build_option_group(
    "juice_fruit",
    "Pilihan Buah",
    [("buah_naga", "Buah Naga"), ("alpukat", "Alpukat"), ("sirsak", "Sirsak")]
)
STEAM_SAUCE_OPTIONS = build_option_group(
    "steam_sauce",
    "Pilihan Saus",
    [("kecap_rica", "Kecap Rica"), ("bawang_putih", "Bawang Putih")]
)
SATE_SAUCE_OPTIONS = build_option_group(
    "sate_sauce",
    "Penyajian Saus",
    [("campur", "Campur Saus"), ("pisah", "Pisah Saus")]
)

OPTION_GROUP_CATALOG = [
    CHICKEN_GEPREK_PART_OPTIONS,
    CHICKEN_STANDARD_PART_OPTIONS,
    GEPREK_RICA_OPTIONS,
    TUNA_COOKING_OPTIONS,
    EGG_STYLE_OPTIONS,
    FRIED_RICE_EGG_OPTIONS,
    BROTH_STYLE_OPTIONS,
    BANANA_TYPE_OPTIONS,
    SERUT_STYLE_OPTIONS,
    BIG_FRIED_STYLE_OPTIONS,
    BIG_FRIED_SIZE_OPTIONS,
    BIG_FRIED_SEASONING_OPTIONS,
    TOPPING_SAUCE_OPTIONS,
    TEMPERATURE_OPTIONS,
    FRUIT_JUICE_OPTIONS,
    STEAM_SAUCE_OPTIONS,
    SATE_SAUCE_OPTIONS
]

MENU_DISPLAY_GROUPS = [
    {
        "id": "ayam_geprek",
        "name": "Ayam Geprek",
        "match_names": ["Ayam Geprek"],
        "option_groups": [SERVING_TYPE_OPTIONS, CHICKEN_GEPREK_PART_OPTIONS, GEPREK_RICA_OPTIONS],
        "combo_keys": ["serving_type"],
        "combo_variants": [
            {"combo": {"serving_type": "pcs"}, "name": "Ayam Geprek", "serving_type": "pcs"},
            {"combo": {"serving_type": "paket"}, "name": "Ayam Geprek", "serving_type": "paket"}
        ]
    },
    {
        "id": "ayam_crispy",
        "name": "Ayam Crispy",
        "match_names": ["Ayam Crispy"],
        "option_groups": [SERVING_TYPE_OPTIONS, CHICKEN_GEPREK_PART_OPTIONS],
        "combo_keys": ["serving_type"],
        "combo_variants": [
            {"combo": {"serving_type": "pcs"}, "name": "Ayam Crispy", "serving_type": "pcs"},
            {"combo": {"serving_type": "paket"}, "name": "Ayam Crispy", "serving_type": "paket"}
        ]
    },
    {
        "id": "ayam_kalasan",
        "name": "Ayam Kalasan",
        "match_names": ["Ayam Kalasan"],
        "option_groups": [SERVING_TYPE_OPTIONS, CHICKEN_STANDARD_PART_OPTIONS],
        "combo_keys": ["serving_type"],
        "combo_variants": [
            {"combo": {"serving_type": "pcs"}, "name": "Ayam Kalasan", "serving_type": "pcs"},
            {"combo": {"serving_type": "paket"}, "name": "Ayam Kalasan", "serving_type": "paket"}
        ]
    },
    {
        "id": "ayam_bakar_rica",
        "name": "Ayam Bakar Rica",
        "match_names": ["Ayam Bakar Rica"],
        "option_groups": [SERVING_TYPE_OPTIONS, CHICKEN_STANDARD_PART_OPTIONS],
        "combo_keys": ["serving_type"],
        "combo_variants": [
            {"combo": {"serving_type": "pcs"}, "name": "Ayam Bakar Rica", "serving_type": "pcs"},
            {"combo": {"serving_type": "paket"}, "name": "Ayam Bakar Rica", "serving_type": "paket"}
        ]
    },
    {
        "id": "ayam_bakar_kecap",
        "name": "Ayam Bakar Kecap",
        "match_names": ["Ayam Bakar Kecap"],
        "option_groups": [SERVING_TYPE_OPTIONS, CHICKEN_STANDARD_PART_OPTIONS],
        "combo_keys": ["serving_type"],
        "combo_variants": [
            {"combo": {"serving_type": "pcs"}, "name": "Ayam Bakar Kecap", "serving_type": "pcs"},
            {"combo": {"serving_type": "paket"}, "name": "Ayam Bakar Kecap", "serving_type": "paket"}
        ]
    },
    {
        "id": "ayam_bakar_iloni",
        "name": "Ayam Bakar Iloni",
        "match_names": ["Ayam Bakar Iloni"],
        "option_groups": [SERVING_TYPE_OPTIONS, CHICKEN_STANDARD_PART_OPTIONS],
        "combo_keys": ["serving_type"],
        "combo_variants": [
            {"combo": {"serving_type": "pcs"}, "name": "Ayam Bakar Iloni", "serving_type": "pcs"},
            {"combo": {"serving_type": "paket"}, "name": "Ayam Bakar Iloni", "serving_type": "paket"}
        ]
    },
    {
        "id": "ayam_serundeng_manna",
        "name": "Ayam Serundeng Manna",
        "match_names": ["Ayam Serundeng Manna"],
        "option_groups": [SERVING_TYPE_OPTIONS, CHICKEN_STANDARD_PART_OPTIONS],
        "combo_keys": ["serving_type"],
        "combo_variants": [
            {"combo": {"serving_type": "pcs"}, "name": "Ayam Serundeng Manna", "serving_type": "pcs"},
            {"combo": {"serving_type": "paket"}, "name": "Ayam Serundeng Manna", "serving_type": "paket"}
        ]
    },
    {
        "id": "dada_tuna",
        "name": "Dada Tuna",
        "match_names": ["Dada Tuna Goreng", "dada tuna bakar"],
        "option_groups": [TUNA_COOKING_OPTIONS],
        "combo_keys": ["cook_style"],
        "combo_variants": [
            {"combo": {"cook_style": "goreng"}, "name": "Dada Tuna Goreng", "serving_type": "pcs"},
            {"combo": {"cook_style": "bakar"}, "name": "dada tuna bakar", "serving_type": "pcs"}
        ]
    },
    {
        "id": "paket_dada_tuna",
        "name": "Paket Dada Tuna",
        "match_names": ["Paket Dada Tuna goreng", "paket dada tuna bakar"],
        "option_groups": [TUNA_COOKING_OPTIONS],
        "combo_keys": ["cook_style"],
        "combo_variants": [
            {"combo": {"cook_style": "goreng"}, "name": "Paket Dada Tuna goreng", "serving_type": "paket"},
            {"combo": {"cook_style": "bakar"}, "name": "paket dada tuna bakar", "serving_type": "paket"}
        ]
    },
    {
        "id": "pisang_goreng",
        "name": "Pisang Goreng",
        "match_names": ["Pisang Goreng Raja", "Pisang Goreng Pagata"],
        "option_groups": [BANANA_TYPE_OPTIONS],
        "combo_keys": ["banana_type"],
        "combo_variants": [
            {"combo": {"banana_type": "raja"}, "name": "Pisang Goreng Raja"},
            {"combo": {"banana_type": "pagata"}, "name": "Pisang Goreng Pagata"}
        ]
    },
    {
        "id": "es_serut_kacang_susu",
        "name": "Es Serut Kacang Susu",
        "match_names": ["Es Serut Kacang Susu Gulmer", "Es Serut Kacang Susu Sirup"],
        "option_groups": [SERUT_STYLE_OPTIONS],
        "combo_keys": ["serut_style"],
        "combo_variants": [
            {"combo": {"serut_style": "gulmer"}, "name": "Es Serut Kacang Susu Gulmer"},
            {"combo": {"serut_style": "sirup"}, "name": "Es Serut Kacang Susu Sirup"}
        ]
    },
    {
        "id": "big_fried_chicken",
        "name": "Big Fried Chicken",
        "match_names": [
            "Big Fried Chicken Original small",
            "Big Fried Chicken Original large",
            "Big Fried Chicken Bubble small",
            "Big Fried Chicken Bubble large"
        ],
        "option_groups": [BIG_FRIED_STYLE_OPTIONS, BIG_FRIED_SIZE_OPTIONS, BIG_FRIED_SEASONING_OPTIONS],
        "combo_keys": ["fried_style", "fried_size"],
        "combo_variants": [
            {"combo": {"fried_style": "ori", "fried_size": "small"}, "name": "Big Fried Chicken Original small"},
            {"combo": {"fried_style": "ori", "fried_size": "large"}, "name": "Big Fried Chicken Original large"},
            {"combo": {"fried_style": "bubble", "fried_size": "small"}, "name": "Big Fried Chicken Bubble small"},
            {"combo": {"fried_style": "bubble", "fried_size": "large"}, "name": "Big Fried Chicken Bubble large"}
        ]
    },
    {
        "id": "big_fried_chicken_topping",
        "name": "Big Fried Chicken + Topping",
        "match_names": [
            "Big Fried Chicken Original small + topping",
            "Big Fried Chicken Original large + topping",
            "Big Fried Chicken Bubble small + topping",
            "Big Fried Chicken Bubble large + topping"
        ],
        "option_groups": [BIG_FRIED_STYLE_OPTIONS, BIG_FRIED_SIZE_OPTIONS, BIG_FRIED_SEASONING_OPTIONS, TOPPING_SAUCE_OPTIONS],
        "combo_keys": ["fried_style", "fried_size"],
        "combo_variants": [
            {"combo": {"fried_style": "ori", "fried_size": "small"}, "name": "Big Fried Chicken Original small + topping"},
            {"combo": {"fried_style": "ori", "fried_size": "large"}, "name": "Big Fried Chicken Original large + topping"},
            {"combo": {"fried_style": "bubble", "fried_size": "small"}, "name": "Big Fried Chicken Bubble small + topping"},
            {"combo": {"fried_style": "bubble", "fried_size": "large"}, "name": "Big Fried Chicken Bubble large + topping"}
        ]
    },
    {
        "id": "ayam_ori_nasi_topping",
        "name": "Ayam Ori + Nasi + Topping",
        "match_names": ["Ayam Ori chilin + nasi + topping"],
        "option_groups": [BIG_FRIED_SEASONING_OPTIONS, TOPPING_SAUCE_OPTIONS],
        "combo_keys": [],
        "combo_variants": [
            {"combo": {}, "name": "Ayam Ori chilin + nasi + topping", "serving_type": "paket"}
        ]
    },
    {
        "id": "ayam_bubble_nasi_topping",
        "name": "Ayam Bubble + Nasi + Topping",
        "match_names": ["Ayam Bubble Chilin + nasi + topping"],
        "option_groups": [BIG_FRIED_SEASONING_OPTIONS, TOPPING_SAUCE_OPTIONS],
        "combo_keys": [],
        "combo_variants": [
            {"combo": {}, "name": "Ayam Bubble Chilin + nasi + topping", "serving_type": "paket"}
        ]
    },
    {
        "id": "chicken_skin_nasi_topping",
        "name": "Chicken Skin + Nasi + Topping",
        "match_names": ["Chicken Skin"],
        "option_groups": [TOPPING_SAUCE_OPTIONS],
        "combo_keys": [],
        "combo_variants": [
            {"combo": {}, "name": "Chicken Skin", "serving_type": "paket"}
        ]
    },
    {
        "id": "teh_tawar",
        "name": "Teh Tawar",
        "match_names": ["Teh Tawar", "Teh Tawar Panas", "Teh Tawar Dingin"],
        "option_groups": [TEMPERATURE_OPTIONS],
        "combo_keys": ["temperature"],
        "combo_variants": [
            {"combo": {"temperature": "ice"}, "name": "Teh Tawar Dingin"},
            {"combo": {"temperature": "hangat"}, "name": "Teh Tawar"},
            {"combo": {"temperature": "hot"}, "name": "Teh Tawar Panas"}
        ]
    },
    {
        "id": "teh_manis",
        "name": "Teh Manis",
        "match_names": ["Teh Manis", "Teh Manis Panas", "Teh Manis Dingin"],
        "option_groups": [TEMPERATURE_OPTIONS],
        "combo_keys": ["temperature"],
        "combo_variants": [
            {"combo": {"temperature": "ice"}, "name": "Teh Manis Dingin"},
            {"combo": {"temperature": "hangat"}, "name": "Teh Manis"},
            {"combo": {"temperature": "hot"}, "name": "Teh Manis Panas"}
        ]
    },
    {
        "id": "kopi_hitam",
        "name": "Kopi Hitam",
        "match_names": ["Kopi Hitam", "Kopi Hitam Panas", "Kopi Hitam Dingin"],
        "option_groups": [TEMPERATURE_OPTIONS],
        "combo_keys": ["temperature"],
        "combo_variants": [
            {"combo": {"temperature": "ice"}, "name": "Kopi Hitam Dingin"},
            {"combo": {"temperature": "hangat"}, "name": "Kopi Hitam"},
            {"combo": {"temperature": "hot"}, "name": "Kopi Hitam Panas"}
        ]
    },
    {
        "id": "kopi_susu",
        "name": "Kopi Susu",
        "match_names": ["Kopi Susu", "Kopi Susu Panas", "Kopi Susu Dingin"],
        "option_groups": [TEMPERATURE_OPTIONS],
        "combo_keys": ["temperature"],
        "combo_variants": [
            {"combo": {"temperature": "ice"}, "name": "Kopi Susu Dingin"},
            {"combo": {"temperature": "hangat"}, "name": "Kopi Susu"},
            {"combo": {"temperature": "hot"}, "name": "Kopi Susu Panas"}
        ]
    },
    {
        "id": "nutrisari",
        "name": "Nutrisari",
        "match_names": ["nutrisari dingin", "nutrisari Dingin", "nutrisari Panas"],
        "option_groups": [TEMPERATURE_OPTIONS],
        "combo_keys": ["temperature"],
        "combo_variants": [
            {"combo": {"temperature": "ice"}, "name": "nutrisari dingin"},
            {"combo": {"temperature": "hangat"}, "name": "nutrisari Panas"},
            {"combo": {"temperature": "hot"}, "name": "nutrisari Panas"}
        ]
    },
    {
        "id": "jeruk_nipis",
        "name": "Jeruk Nipis",
        "match_names": ["Jeruk Nipis", "Jeruk Nipis Panas", "Jeruk Nipis Dingin"],
        "option_groups": [TEMPERATURE_OPTIONS],
        "combo_keys": ["temperature"],
        "combo_variants": [
            {"combo": {"temperature": "ice"}, "name": "Jeruk Nipis Dingin"},
            {"combo": {"temperature": "hangat"}, "name": "Jeruk Nipis"},
            {"combo": {"temperature": "hot"}, "name": "Jeruk Nipis Panas"}
        ]
    },
    {
        "id": "jeruk_manis",
        "name": "Jeruk Manis",
        "match_names": ["Jeruk Manis", "Jeruk Manis Panas", "Jeruk Manis Dingin"],
        "option_groups": [TEMPERATURE_OPTIONS],
        "combo_keys": ["temperature"],
        "combo_variants": [
            {"combo": {"temperature": "ice"}, "name": "Jeruk Manis Dingin"},
            {"combo": {"temperature": "hangat"}, "name": "Jeruk Manis"},
            {"combo": {"temperature": "hot"}, "name": "Jeruk Manis Panas"}
        ]
    },
    {
        "id": "lemon_tea",
        "name": "Lemon Tea",
        "match_names": ["Lemon Tea", "Lemon Tea Panas", "Lemon Tea Dingin"],
        "option_groups": [TEMPERATURE_OPTIONS],
        "combo_keys": ["temperature"],
        "combo_variants": [
            {"combo": {"temperature": "ice"}, "name": "Lemon Tea Dingin"},
            {"combo": {"temperature": "hangat"}, "name": "Lemon Tea"},
            {"combo": {"temperature": "hot"}, "name": "Lemon Tea Panas"}
        ]
    },
    {
        "id": "jus_buah",
        "name": "Jus Buah",
        "match_names": ["Jus Buah Naga", "Jus Buah Sirsak", "Jus Buah Avocado"],
        "option_groups": [FRUIT_JUICE_OPTIONS],
        "combo_keys": ["juice_fruit"],
        "combo_variants": [
            {"combo": {"juice_fruit": "buah_naga"}, "name": "Jus Buah Naga"},
            {"combo": {"juice_fruit": "alpukat"}, "name": "Jus Buah Avocado"},
            {"combo": {"juice_fruit": "sirsak"}, "name": "Jus Buah Sirsak"}
        ]
    }
]

MENU_SINGLE_OPTION_CONFIGS = {
    "ayam geprek": {
        "display_name": "Ayam Geprek",
        "option_groups": [GEPREK_RICA_OPTIONS]
    },
    "ayam crispy": {
        "display_name": "Ayam Crispy",
        "option_groups": []
    },
    "ayam kalasan": {
        "display_name": "Ayam Kalasan",
        "option_groups": []
    },
    "ayam bakar rica": {
        "display_name": "Ayam Bakar Rica",
        "option_groups": []
    },
    "ayam bakar kecap": {
        "display_name": "Ayam Bakar Kecap",
        "option_groups": []
    },
    "ayam bakar iloni": {
        "display_name": "Ayam Bakar Iloni",
        "option_groups": []
    },
    "ayam serundeng manna": {
        "display_name": "Ayam Serundeng Manna",
        "option_groups": []
    },
    "telur": {
        "display_name": "Telur",
        "option_groups": [EGG_STYLE_OPTIONS]
    },
    "nasi goreng kampung": {
        "display_name": "Nasi Goreng Kampung",
        "option_groups": [FRIED_RICE_EGG_OPTIONS]
    },
    "nasi goreng sagela": {
        "display_name": "Nasi Goreng Sagela",
        "option_groups": [FRIED_RICE_EGG_OPTIONS]
    },
    "nasi goreng spesial": {
        "display_name": "Nasi Goreng Spesial",
        "option_groups": [FRIED_RICE_EGG_OPTIONS]
    },
    "nasi goreng pete": {
        "display_name": "Nasi Goreng Pete",
        "option_groups": [FRIED_RICE_EGG_OPTIONS]
    },
    "mie titi": {
        "display_name": "Mie Titi",
        "option_groups": [BROTH_STYLE_OPTIONS]
    },
    "nila steam kecap rica": {
        "display_name": "Nila Steam",
        "option_groups": [STEAM_SAUCE_OPTIONS]
    },
    "sate ayam (6 tusuk)": {
        "display_name": "Sate Ayam (6 Tusuk)",
        "option_groups": [SATE_SAUCE_OPTIONS]
    },
    "sate tuna (5 tusuk)": {
        "display_name": "Sate Tuna (5 Tusuk)",
        "option_groups": [SATE_SAUCE_OPTIONS]
    },
    "sate daging (5 tusuk)": {
        "display_name": "Sate Daging (5 Tusuk)",
        "option_groups": [SATE_SAUCE_OPTIONS]
    },
    "big fried chicken original small": {
        "display_name": "Big Fried Chicken Original small",
        "option_groups": [BIG_FRIED_SEASONING_OPTIONS]
    },
    "big fried chicken original large": {
        "display_name": "Big Fried Chicken Original large",
        "option_groups": [BIG_FRIED_SEASONING_OPTIONS]
    },
    "big fried chicken bubble small": {
        "display_name": "Big Fried Chicken Bubble small",
        "option_groups": [BIG_FRIED_SEASONING_OPTIONS]
    },
    "big fried chicken bubble large": {
        "display_name": "Big Fried Chicken Bubble large",
        "option_groups": [BIG_FRIED_SEASONING_OPTIONS]
    },
    "big fried chicken original small + topping": {
        "display_name": "Big Fried Chicken Original small + topping",
        "option_groups": [BIG_FRIED_SEASONING_OPTIONS, TOPPING_SAUCE_OPTIONS]
    },
    "big fried chicken original large + topping": {
        "display_name": "Big Fried Chicken Original large + topping",
        "option_groups": [BIG_FRIED_SEASONING_OPTIONS, TOPPING_SAUCE_OPTIONS]
    },
    "big fried chicken bubble small + topping": {
        "display_name": "Big Fried Chicken Bubble small + topping",
        "option_groups": [BIG_FRIED_SEASONING_OPTIONS, TOPPING_SAUCE_OPTIONS]
    },
    "big fried chicken bubble large + topping": {
        "display_name": "Big Fried Chicken Bubble large + topping",
        "option_groups": [BIG_FRIED_SEASONING_OPTIONS, TOPPING_SAUCE_OPTIONS]
    },
    "ayam ori chilin + nasi + topping": {
        "display_name": "Ayam Ori chilin + nasi + topping",
        "option_groups": [BIG_FRIED_SEASONING_OPTIONS, TOPPING_SAUCE_OPTIONS]
    },
    "ayam bubble chilin + nasi + topping": {
        "display_name": "Ayam Bubble Chilin + nasi + topping",
        "option_groups": [BIG_FRIED_SEASONING_OPTIONS, TOPPING_SAUCE_OPTIONS]
    },
    "chicken skin": {
        "display_name": "Chicken Skin",
        "option_groups": [TOPPING_SAUCE_OPTIONS]
    }
}

OPTION_STOCK_LABELS = OrderedDict([
    ("seasoning", OrderedDict([
        ("balado", "Balado"),
        ("keju", "Keju"),
        ("bbq", "BBQ"),
        ("jagung_bakar", "Jagung Bakar"),
        ("sapi_panggang", "Sapi Panggang"),
        ("extra_hot", "Extra Hot")
    ])),
    ("topping_sauce", OrderedDict([
        ("saus_keju", "Saus Keju"),
        ("saus_mayo", "Saus Mayo"),
        ("saus_bbq", "Saus BBQ"),
        ("saus_pedas_manis", "Saus Pedas Manis"),
        ("saus_rica_bawang", "Saus Rica Bawang")
    ]))
])


# ================= DATABASE CONNECTION =================

def get_db_connection():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="moca_restaurant"
    )


def get_restaurant_whatsapp_link(message="Halo Manna Bakery and Cafe, saya ingin tanya reservasi meja."):
    return f"https://wa.me/{RESTAURANT_PROFILE['whatsapp_number']}?text={quote(message)}"


def build_public_booking_description(whatsapp_number, seating_preference, notes):
    parts = ["Sumber: Website Booking"]

    if whatsapp_number:
        parts.append(f"WhatsApp: {whatsapp_number}")

    if seating_preference:
        parts.append(f"Preferensi meja: {seating_preference}")

    if notes:
        parts.append(f"Catatan: {notes}")

    return " | ".join(parts)


def dish_description_sql(item_alias="ri"):
    return f"""
        COALESCE(
            NULLIF(TRIM(REPLACE({item_alias}.dish_description,'  ',' ')),''),
            CASE
                WHEN LOWER(TRIM(COALESCE({item_alias}.special_request,''))) IN ('', 'no_special', 'with_special') THEN NULL
                ELSE NULLIF(TRIM(REPLACE({item_alias}.special_request,'  ',' ')),'')
            END
        )
    """


def format_fish_info(row):
    fish_type = (row.get("fish_type") or "").strip()
    fish_size = (row.get("fish_size") or "").strip()
    fish_weight = row.get("fish_weight") or 0
    fish_weight_unit = row.get("fish_weight_unit") or row.get("weight_unit") or "ons"

    parts = []

    if fish_size:
        parts.append(f"Nila {fish_size.capitalize()}")
    elif fish_type and fish_type.lower() != "none":
        parts.append(fish_type.capitalize())

    if fish_weight and fish_weight > 0:
        parts.append(format_weight_display(fish_weight, fish_weight_unit))

    return " - ".join(parts)


def format_currency(value):
    if value in (None, ""):
        return None

    try:
        return f"{int(value):,}".replace(",", ".")
    except (ValueError, TypeError):
        return str(value)


def format_menu_label(menu_name, serving_type):
    name = (menu_name or "").strip()
    serving = (serving_type or "").strip().lower()

    if not name:
        return "-"

    normalized_name = name.lower()
    serving_map = {
        "paket": "PAKET",
        "package": "PAKET",
        "pcs": "PCS",
        "piece": "PCS",
        "menu_piece": "PCS",
        "porsi": "PORSI",
        "portion": "PORSI"
    }
    serving_label = serving_map.get(serving)

    if serving_label:
        suffix = f"({serving_label})"
        if suffix.lower() in normalized_name:
            return name
        return f"{name} {suffix}"

    return name


def normalize_text(value):
    return re.sub(r"\s+", " ", (value or "").strip().lower())


def normalize_serving_type(serving_type):
    serving = normalize_text(serving_type)
    serving_map = {
        "paket": "paket",
        "package": "paket",
        "pcs": "pcs",
        "piece": "pcs",
        "menu_piece": "pcs",
        "porsi": "porsi",
        "portion": "porsi"
    }
    return serving_map.get(serving, serving)


def serialize_combo_key(combo_values, combo_keys):
    if not combo_keys:
        return "__default__"
    return "|".join(
        f"{key}={normalize_text(combo_values.get(key))}"
        for key in combo_keys
    )


def normalize_menu_options_payload(menu_options_json):
    payload = {}

    if not menu_options_json:
        return payload

    if isinstance(menu_options_json, dict):
        payload = dict(menu_options_json)
    else:
        try:
            payload = json.loads(menu_options_json)
        except (TypeError, ValueError, json.JSONDecodeError):
            return {}

    if not isinstance(payload, dict):
        return {}

    selected_options = payload.get("selected_options")
    if isinstance(selected_options, dict):
        selected_options = dict(selected_options)
        selected_options.pop("serving_type", None)
        selected_options.pop("chicken_part", None)
        payload["selected_options"] = selected_options

    return payload


def normalize_option_value_list(raw_value):
    if raw_value is None:
        return []

    if isinstance(raw_value, (list, tuple, set)):
        values = raw_value
    else:
        values = [raw_value]

    normalized_values = []
    for value in values:
        if value is None:
            continue
        text = str(value).strip()
        if text:
            normalized_values.append(text)
    return normalized_values


def extract_selected_option_values(menu_options_json):
    payload = normalize_menu_options_payload(menu_options_json)
    selected_options = payload.get("selected_options") or {}
    option_values = {}

    for option_key, option in selected_options.items():
        raw_value = option.get("value")
        if isinstance(raw_value, (list, tuple, set)):
            option_values[option_key] = normalize_option_value_list(raw_value)
        else:
            normalized_values = normalize_option_value_list(raw_value)
            option_values[option_key] = normalized_values[0] if normalized_values else ""

    return option_values


def get_option_group_metadata():
    metadata = {}
    for group in OPTION_GROUP_CATALOG:
        bucket = metadata.setdefault(group["key"], {
            "label": group["label"],
            "multiple": group.get("multiple", False),
            "choices": {}
        })
        for choice in group.get("choices", []):
            bucket["choices"][choice["value"]] = choice["label"]
    return metadata


OPTION_GROUP_METADATA = get_option_group_metadata()


def build_selected_option_entry(option_key, raw_value):
    meta = OPTION_GROUP_METADATA.get(option_key, {})
    label = meta.get("label") or option_key.replace("_", " ").title()
    choices = meta.get("choices", {})
    multiple = meta.get("multiple", False)

    normalized_values = normalize_option_value_list(raw_value)
    if not normalized_values:
        return None

    display_values = [choices.get(value, value) for value in normalized_values]
    if multiple or len(normalized_values) > 1:
        return {
            "label": label,
            "value": normalized_values,
            "display": display_values
        }

    return {
        "label": label,
        "value": normalized_values[0],
        "display": display_values[0]
    }


def build_menu_options_payload_from_form(form_data, menu_catalog, display_menu_id, fallback_payload=None):
    selected_display_menu_id = str(display_menu_id or "").strip()
    if not selected_display_menu_id:
        return fallback_payload

    selected_menu = next(
        (menu for menu in menu_catalog if str(menu.get("id")) == selected_display_menu_id),
        None
    )
    if not selected_menu or not selected_menu.get("option_groups"):
        return fallback_payload

    selected_options = OrderedDict()
    for group in selected_menu.get("option_groups", []):
        option_key = group.get("key")
        field_name = f"option__{option_key}"

        if group.get("multiple"):
            raw_value = form_data.getlist(field_name)
        else:
            raw_value = form_data.get(field_name)

        option_entry = build_selected_option_entry(option_key, raw_value)
        if option_entry:
            selected_options[option_key] = option_entry

    if not selected_options:
        return fallback_payload

    return json.dumps({
        "display_menu_id": selected_menu.get("id"),
        "display_name": selected_menu.get("name"),
        "selected_options": selected_options
    })


def infer_selected_options_from_row(row):
    inferred_options = OrderedDict()
    name_key = normalize_text(row.get("name"))

    if "big fried chicken original" in name_key or "ayam ori chilin" in name_key:
        option_entry = build_selected_option_entry("fried_style", "ori")
        if option_entry:
            inferred_options["fried_style"] = option_entry
    elif "big fried chicken bubble" in name_key or "ayam bubble chilin" in name_key:
        option_entry = build_selected_option_entry("fried_style", "bubble")
        if option_entry:
            inferred_options["fried_style"] = option_entry

    if " small" in name_key:
        option_entry = build_selected_option_entry("fried_size", "small")
        if option_entry:
            inferred_options["fried_size"] = option_entry
    elif " large" in name_key:
        option_entry = build_selected_option_entry("fried_size", "large")
        if option_entry:
            inferred_options["fried_size"] = option_entry

    if " dingin" in name_key or " ice" in name_key:
        option_entry = build_selected_option_entry("temperature", "ice")
        if option_entry:
            inferred_options["temperature"] = option_entry
    elif " panas" in name_key or name_key.endswith(" hot"):
        option_entry = build_selected_option_entry("temperature", "hot")
        if option_entry:
            inferred_options["temperature"] = option_entry
    elif name_key in {
        "teh tawar", "teh manis", "kopi hitam", "kopi susu", "nutrisari panas",
        "jeruk nipis", "jeruk manis", "lemon tea"
    }:
        option_entry = build_selected_option_entry("temperature", "hangat")
        if option_entry:
            inferred_options["temperature"] = option_entry

    if "dada tuna bakar" in name_key or "paket dada tuna bakar" in name_key:
        option_entry = build_selected_option_entry("cook_style", "bakar")
        if option_entry:
            inferred_options["cook_style"] = option_entry
    elif "dada tuna goreng" in name_key or "paket dada tuna goreng" in name_key:
        option_entry = build_selected_option_entry("cook_style", "goreng")
        if option_entry:
            inferred_options["cook_style"] = option_entry

    if "pisang goreng raja" in name_key:
        option_entry = build_selected_option_entry("banana_type", "raja")
        if option_entry:
            inferred_options["banana_type"] = option_entry
    elif "pisang goreng pagata" in name_key:
        option_entry = build_selected_option_entry("banana_type", "pagata")
        if option_entry:
            inferred_options["banana_type"] = option_entry

    if "gulmer" in name_key:
        option_entry = build_selected_option_entry("serut_style", "gulmer")
        if option_entry:
            inferred_options["serut_style"] = option_entry
    elif "sirup" in name_key:
        option_entry = build_selected_option_entry("serut_style", "sirup")
        if option_entry:
            inferred_options["serut_style"] = option_entry

    if "buah naga" in name_key:
        option_entry = build_selected_option_entry("juice_fruit", "buah_naga")
        if option_entry:
            inferred_options["juice_fruit"] = option_entry
    elif "avocado" in name_key or "alpukat" in name_key:
        option_entry = build_selected_option_entry("juice_fruit", "alpukat")
        if option_entry:
            inferred_options["juice_fruit"] = option_entry
    elif "sirsak" in name_key:
        option_entry = build_selected_option_entry("juice_fruit", "sirsak")
        if option_entry:
            inferred_options["juice_fruit"] = option_entry

    return inferred_options


def get_effective_selected_options(row):
    payload = normalize_menu_options_payload(row.get("menu_options_json"))
    selected_options = OrderedDict(payload.get("selected_options") or {})
    inferred_options = infer_selected_options_from_row(row)

    for option_key, option in inferred_options.items():
        selected_options.setdefault(option_key, option)

    return selected_options


def resolve_menu_submission(menu_catalog, submitted_menu_id, display_menu_id=None, menu_options_json=None):
    payload = normalize_menu_options_payload(menu_options_json)
    effective_display_menu_id = str(
        display_menu_id or payload.get("display_menu_id") or ""
    ).strip()

    if effective_display_menu_id:
        selected_menu = next(
            (menu for menu in menu_catalog if str(menu.get("id")) == effective_display_menu_id),
            None
        )
        if selected_menu:
            if selected_menu.get("combo_map"):
                option_values = extract_selected_option_values(menu_options_json)
                combo_key = serialize_combo_key(
                    option_values,
                    selected_menu.get("combo_keys", [])
                )
                resolved_choice = (
                    selected_menu.get("combo_map", {}).get(combo_key)
                    or next(iter(selected_menu.get("combo_map", {}).values()), None)
                )
                if resolved_choice and resolved_choice.get("menu_id") is not None:
                    return int(resolved_choice["menu_id"])

            if selected_menu.get("menu_id") is not None:
                return int(selected_menu["menu_id"])

    submitted_menu_id = str(submitted_menu_id or "").strip()
    if submitted_menu_id == "":
        return None

    if submitted_menu_id.lstrip("-").isdigit():
        return int(submitted_menu_id)

    return None


def format_option_display_value(raw_value):
    return ", ".join(normalize_option_value_list(raw_value))


def build_option_summary_from_payload(menu_options_payload):
    payload = normalize_menu_options_payload(menu_options_payload)
    selected_options = payload.get("selected_options") or {}
    summary_parts = []

    for option in selected_options.values():
        label = (option.get("label") or "").strip()
        display = format_option_display_value(option.get("display"))
        if label and display:
            summary_parts.append(f"{label}: {display}")

    return " | ".join(summary_parts)


def build_option_summary_from_row(row):
    summary_parts = []

    for option in (row.get("effective_selected_options") or get_effective_selected_options(row)).values():
        label = (option.get("label") or "").strip()
        display = format_option_display_value(option.get("display"))
        if label and display:
            summary_parts.append(f"{label}: {display}")

    return " | ".join(summary_parts)


def get_payload_display_name(menu_options_payload, fallback_name, fallback_serving_type=None):
    payload = normalize_menu_options_payload(menu_options_payload)
    payload_name = (payload.get("display_name") or "").strip()
    if payload_name:
        return payload_name
    return normalize_base_menu_name(fallback_name, fallback_serving_type)


def build_menu_display_note(menu_options_payload, custom_note=None):
    option_summary = build_option_summary_from_payload(menu_options_payload)
    note_parts = []

    if option_summary:
        note_parts.append(option_summary)

    if custom_note:
        note_parts.append(custom_note.strip())

    return " | ".join(part for part in note_parts if part)


def build_menu_display_note_from_row(row):
    note_parts = []
    option_summary = build_option_summary_from_row(row)
    custom_note = (row.get("dish_description") or "").strip()

    if option_summary:
        note_parts.append(option_summary)

    if custom_note:
        note_parts.append(custom_note)

    return " | ".join(part for part in note_parts if part)


def get_latest_menu_status_map(cursor, selected_date):
    cursor.execute("""
    SELECT menu_id, status, stock_date, id
    FROM daily_menu_stock
    WHERE stock_date <= %s
    ORDER BY stock_date DESC, id DESC
    """,(selected_date,))

    status_map = {}
    for row in cursor.fetchall():
        menu_id = row_value(row, "menu_id", 0)
        if menu_id not in status_map:
            status_map[menu_id] = row
    return status_map


def get_latest_nila_status_map(cursor, selected_date):
    cursor.execute("""
    SELECT id, size_category, status, stock_date
    FROM fish_stock
    WHERE fish_type_id = 4
    AND size_category IS NOT NULL
    AND stock_date <= %s
    ORDER BY stock_date DESC, id DESC
    """,(selected_date,))

    nila_map = {}
    for row in cursor.fetchall():
        size_key = row_value(row, "size_category", 1)
        if size_key not in nila_map:
            nila_map[size_key] = row
    return nila_map


def get_latest_daily_item_stock_rows(cursor, menu_ids, selected_date):
    if not menu_ids:
        return []

    placeholders = ",".join(["%s"] * len(menu_ids))
    cursor.execute(
        f"""
        SELECT
            ds.id,
            ds.menu_id,
            ds.weight_ons,
            COALESCE(ds.weight_unit,'ons') AS weight_unit,
            ds.available_qty,
            ds.status,
            ds.stock_date,
            m.name
        FROM daily_item_stock ds
        JOIN menus m ON m.id = ds.menu_id
        WHERE ds.menu_id IN ({placeholders})
        AND ds.stock_date <= %s
        ORDER BY ds.stock_date DESC, ds.id DESC
        """,
        list(menu_ids) + [selected_date]
    )

    latest_rows = OrderedDict()
    for row in cursor.fetchall():
        row_key = (
            row_value(row, "menu_id", 1),
            float(row_value(row, "weight_ons", 2, 0) or 0)
        )
        if row_key not in latest_rows:
            latest_rows[row_key] = row
    return list(latest_rows.values())


def get_latest_option_stock_map(cursor, selected_date):
    cursor.execute("""
    SELECT
        option_key,
        option_value,
        status,
        stock_date,
        id
    FROM menu_option_stock
    WHERE stock_date <= %s
    ORDER BY stock_date DESC, id DESC
    """,(selected_date,))

    option_map = {}
    for row in cursor.fetchall():
        option_key = row_value(row, "option_key", 0)
        option_value = row_value(row, "option_value", 1)
        pair_key = (option_key, option_value)
        if pair_key not in option_map:
            option_map[pair_key] = row_value(row, "status", 2, "ready")
    return option_map


def get_latest_sea_fish_stock_rows(cursor, selected_date):
    cursor.execute("""
    SELECT
        fs.id,
        fs.fish_type_id,
        ft.name,
        fs.weight_ons,
        COALESCE(fs.weight_unit,'ons') AS weight_unit,
        fs.fish_count,
        fs.status,
        fs.stock_date
    FROM fish_stock fs
    JOIN fish_types ft ON ft.id = fs.fish_type_id
    WHERE fs.stock_date <= %s
    AND ft.fish_category = 'sea'
    ORDER BY fs.stock_date DESC, fs.id DESC
    """,(selected_date,))

    latest_rows = OrderedDict()
    for row in cursor.fetchall():
        row_key = (
            row_value(row, "fish_type_id", 1),
            float(row_value(row, "weight_ons", 3, 0) or 0)
        )
        if row_key not in latest_rows:
            latest_rows[row_key] = row

    return list(latest_rows.values())


def apply_option_stock_to_catalog(menu_catalog, option_stock_map):
    for menu in menu_catalog:
        updated_groups = []
        for group in menu.get("option_groups", []):
            group_copy = dict(group)
            updated_choices = []
            for choice in group.get("choices", []):
                choice_copy = dict(choice)
                choice_status = option_stock_map.get((group.get("key"), choice.get("value")), "ready")
                choice_copy["status"] = choice_status
                updated_choices.append(choice_copy)
            group_copy["choices"] = updated_choices
            updated_groups.append(group_copy)
        menu["option_groups"] = updated_groups
    return menu_catalog


def validate_menu_option_stock(cursor, reservation_date, menu_options_json):
    payload = normalize_menu_options_payload(menu_options_json)
    selected_options = payload.get("selected_options") or {}
    if not selected_options:
        return True, None

    option_status_map = get_latest_option_stock_map(cursor, reservation_date)
    for option_key, option in selected_options.items():
        selected_values = normalize_option_value_list(option.get("value"))
        if not selected_values:
            continue

        selected_displays = normalize_option_value_list(option.get("display"))
        for index, selected_value in enumerate(selected_values):
            current_status = option_status_map.get((option_key, selected_value), "ready")
            if current_status != "ready":
                option_label = option.get("label") or option_key
                option_display = (
                    selected_displays[index]
                    if index < len(selected_displays)
                    else selected_value
                )
                return False, f"Pilihan {option_label} - {option_display} sedang not ready."

    return True, None


def log_stock_history(cursor, stock_scope, target_name, previous_value, new_value, actor=None, notes=None):
    if str(previous_value) == str(new_value):
        return

    cursor.execute("""
        INSERT INTO stock_change_log
        (stock_scope, target_name, previous_value, new_value, actor_name, notes)
        VALUES (%s,%s,%s,%s,%s,%s)
    """,(
        stock_scope,
        target_name,
        None if previous_value is None else str(previous_value),
        None if new_value is None else str(new_value),
        actor or "system",
        notes
    ))


def log_reservation_history(cursor, reservation_id, action_type, change_scope, summary, actor=None, reservation_item_id=None):
    cursor.execute("""
        INSERT INTO reservation_change_log
        (reservation_id, reservation_item_id, action_type, change_scope, summary, actor_name)
        VALUES (%s,%s,%s,%s,%s,%s)
    """,(
        reservation_id,
        reservation_item_id,
        action_type,
        change_scope,
        summary,
        actor or "system"
    ))


def normalize_history_note(note):
    cleaned_note = re.sub(r"\s+", " ", (note or "").strip())
    if not cleaned_note:
        return None

    if cleaned_note.rstrip(".").lower() == "tanpa keterangan":
        return None

    return cleaned_note


def get_menu_label_for_history(cursor, menu_id, menu_options_json=None, fallback_name=None, fallback_serving_type=None):
    if fallback_name:
        return get_payload_display_name(menu_options_json, fallback_name, fallback_serving_type)

    payload_name = (normalize_menu_options_payload(menu_options_json).get("display_name") or "").strip()
    if payload_name:
        return payload_name

    if menu_id in (None, ""):
        return "-"

    try:
        normalized_menu_id = int(menu_id)
    except (TypeError, ValueError):
        return str(menu_id)

    cursor.execute("""
        SELECT name, serving_type
        FROM menus
        WHERE id = %s
        LIMIT 1
    """,(normalized_menu_id,))
    menu_row = cursor.fetchone()

    if not menu_row:
        return f"Menu #{normalized_menu_id}"

    return get_payload_display_name(
        menu_options_json,
        menu_row.get("name"),
        menu_row.get("serving_type")
    )


def build_menu_history_summary(action_type, menu_label, qty=None, note=None, previous_menu_label=None, previous_qty=None):
    normalized_action = (action_type or "").strip().lower()
    resolved_menu_label = (menu_label or "menu").strip()
    normalized_note = normalize_history_note(note)

    if normalized_action == "create":
        parts = [f"Menu ditambahkan: {resolved_menu_label}"]
        if qty not in (None, ""):
            parts.append(f"jumlah {qty}")
        summary = ", ".join(parts) + "."
    elif normalized_action == "update":
        if previous_menu_label and previous_menu_label.strip() and previous_menu_label.strip() != resolved_menu_label:
            summary = f"Menu diubah dari {previous_menu_label.strip()} menjadi {resolved_menu_label}"
        else:
            summary = f"Menu diperbarui: {resolved_menu_label}"

        if previous_qty not in (None, "") and qty not in (None, "") and str(previous_qty) != str(qty):
            summary += f", jumlah {previous_qty} menjadi {qty}"
        elif qty not in (None, ""):
            summary += f", jumlah {qty}"

        summary += "."
    elif normalized_action == "delete":
        summary = f"Menu dihapus: {resolved_menu_label}"
        if qty not in (None, ""):
            summary += f", jumlah {qty}"
        summary += "."
    else:
        summary = resolved_menu_label
        if qty not in (None, ""):
            summary += f", jumlah {qty}"
        summary += "."

    if normalized_note:
        summary += f" {normalized_note}"

    return summary


def get_history_action_label(action_type):
    return {
        "create": "Ditambahkan",
        "update": "Diubah",
        "delete": "Dihapus"
    }.get((action_type or "").strip().lower(), (action_type or "-").strip() or "-")


def get_history_scope_label(change_scope):
    return {
        "reservation": "Reservasi",
        "menu": "Menu"
    }.get((change_scope or "").strip().lower(), (change_scope or "-").strip() or "-")


def extract_menu_ids_from_history_summary(summary):
    return {
        int(menu_id)
        for menu_id in re.findall(r"menu\s+#(\d+)", summary or "", flags=re.IGNORECASE)
    }


def get_history_menu_name_map(cursor, history_rows):
    menu_ids = sorted({
        menu_id
        for row in history_rows
        for menu_id in extract_menu_ids_from_history_summary(row.get("summary"))
    })
    if not menu_ids:
        return {}

    placeholders = ",".join(["%s"] * len(menu_ids))
    cursor.execute(f"""
        SELECT id, name, serving_type
        FROM menus
        WHERE id IN ({placeholders})
    """, tuple(menu_ids))

    return {
        row["id"]: format_menu_label(row.get("name"), row.get("serving_type"))
        for row in cursor.fetchall()
    }


def humanize_reservation_history_summary(row, menu_name_map):
    summary = re.sub(r"\s+", " ", (row.get("summary") or "").strip())
    if not summary:
        return "-"

    action_type = (row.get("action_type") or "").strip().lower()
    change_scope = (row.get("change_scope") or "").strip().lower()

    if change_scope == "menu":
        create_match = re.match(
            r"Tambah menu #(?P<menu_id>\d+) qty (?P<qty>\d+)\.\s*(?P<note>.*)$",
            summary,
            flags=re.IGNORECASE
        )
        if action_type == "create" and create_match:
            menu_id = int(create_match.group("menu_id"))
            menu_label = menu_name_map.get(menu_id) or f"Menu #{menu_id}"
            return build_menu_history_summary(
                "create",
                menu_label,
                qty=create_match.group("qty"),
                note=create_match.group("note")
            )

        update_match = re.match(
            r"Menu item #(?P<item_id>\d+) diubah ke menu #(?P<menu_id>\d+) qty (?P<qty>\d+)\.\s*(?P<note>.*)$",
            summary,
            flags=re.IGNORECASE
        )
        if action_type == "update" and update_match:
            menu_id = int(update_match.group("menu_id"))
            menu_label = menu_name_map.get(menu_id) or f"Menu #{menu_id}"
            return build_menu_history_summary(
                "update",
                menu_label,
                qty=update_match.group("qty"),
                note=update_match.group("note")
            )

        delete_match = re.match(
            r"Menu (?P<menu_name>.+?) qty (?P<qty>\d+) dihapus\.\s*$",
            summary,
            flags=re.IGNORECASE
        )
        if action_type == "delete" and delete_match:
            return build_menu_history_summary(
                "delete",
                delete_match.group("menu_name"),
                qty=delete_match.group("qty")
            )

    readable_summary = re.sub(r"Menu item #\d+\s+", "Item menu ", summary, flags=re.IGNORECASE)
    readable_summary = re.sub(r"\bqty\b", "jumlah", readable_summary, flags=re.IGNORECASE)
    readable_summary = re.sub(
        r"menu\s+#(\d+)",
        lambda match: f"menu {menu_name_map.get(int(match.group(1))) or f'#{match.group(1)}'}",
        readable_summary,
        flags=re.IGNORECASE
    )

    return readable_summary


def prepare_reservation_history_rows(cursor, history_rows):
    menu_name_map = get_history_menu_name_map(cursor, history_rows)

    for row in history_rows:
        row["display_action"] = get_history_action_label(row.get("action_type"))
        row["display_scope"] = get_history_scope_label(row.get("change_scope"))
        row["display_summary"] = humanize_reservation_history_summary(row, menu_name_map)
        row["display_reservation_id"] = f"Reservasi #{row['reservation_id']}" if row.get("reservation_id") else "-"
        row["display_item_id"] = f"Item menu #{row['reservation_item_id']}" if row.get("reservation_item_id") else "-"

    return history_rows


def ensure_menu_catalog_updates(cursor):
    cursor.execute("""
        UPDATE menus
        SET divisi = 'local'
        WHERE LOWER(name) = 'udang sambal pete'
        AND divisi <> 'local'
    """)
    cursor.execute("""
        UPDATE menus
        SET divisi = 'seafood'
        WHERE LOWER(name) = 'paket tuna goreng rica'
        AND divisi <> 'seafood'
    """)
    cursor.execute("""
        INSERT INTO menus (name, category, serving_type, stock_type, divisi, price)
        SELECT %s, %s, %s, %s, %s, %s
        FROM DUAL
        WHERE NOT EXISTS (
            SELECT 1
            FROM menus
            WHERE LOWER(name) = %s
            AND serving_type = %s
        )
    """, (
        "Paket Tuna Goreng Rica",
        "ikan tuna",
        "paket",
        "normal",
        "seafood",
        35000,
        "paket tuna goreng rica",
        "paket"
    ))
    cursor.execute("""
        INSERT INTO menus (name, category, serving_type, stock_type, divisi, price)
        SELECT %s, %s, %s, %s, %s, %s
        FROM DUAL
        WHERE NOT EXISTS (
            SELECT 1
            FROM menus
            WHERE LOWER(name) = %s
            AND serving_type = %s
        )
    """, (
        "Tahu/Tempe Goreng Crispy",
        "side",
        "porsi",
        "normal",
        "seafood",
        20000,
        "tahu/tempe goreng crispy",
        "porsi"
    ))


def build_display_menu_catalog(raw_menus):
    catalog = []

    for raw_menu in raw_menus:
        name_key = normalize_text(raw_menu.get("name"))
        single_config = MENU_SINGLE_OPTION_CONFIGS.get(name_key, {})
        catalog.append({
            "id": f"raw::{raw_menu.get('id')}",
            "name": single_config.get("display_name") or raw_menu.get("name"),
            "category": raw_menu.get("category"),
            "divisi": raw_menu.get("divisi"),
            "stock_type": raw_menu.get("stock_type"),
            "stock_source": raw_menu.get("stock_source") or raw_menu.get("stock_type"),
            "status": raw_menu.get("status") or "ready",
            "option_groups": single_config.get("option_groups", []),
            "combo_keys": [],
            "combo_map": {},
            "menu_id": raw_menu.get("id"),
            "actual_menu_name": raw_menu.get("name"),
            "serving_type": raw_menu.get("serving_type")
        })

    return sorted(catalog, key=lambda item: normalize_text(item.get("name")))


def parse_table_filters(requested_tables, use_table_filters=False):
    normalized_tables = {
        normalize_text(table_key)
        for table_key in requested_tables
        if (table_key or "").strip()
    }

    if not normalized_tables:
        if use_table_filters:
            return {"menu_total"}
        return set(EXPORT_TABLE_KEYS)

    selected_tables = normalized_tables & EXPORT_TABLE_KEYS
    return selected_tables or set(EXPORT_TABLE_KEYS)


def get_default_dashboard_column_keys(table_key):
    return [
        column["key"]
        for column in DASHBOARD_TABLE_CONFIGS.get(table_key, {}).get("columns", [])
    ]


def parse_dashboard_column_filters(request_args, selected_tables):
    selected_columns = {}

    for table_key in DASHBOARD_TABLE_CONFIGS:
        if table_key not in selected_tables:
            continue

        available_columns = DASHBOARD_TABLE_CONFIGS.get(table_key, {}).get("columns", [])
        available_column_keys = {
            normalize_text(column.get("key"))
            for column in available_columns
        }
        requested_keys = [
            normalize_text(column_key)
            for column_key in request_args.getlist(f"columns_{table_key}")
            if (column_key or "").strip()
        ]

        if not requested_keys:
            selected_columns[table_key] = get_default_dashboard_column_keys(table_key)
            continue

        filtered_keys = [
            column["key"]
            for column in available_columns
            if normalize_text(column.get("key")) in available_column_keys
            and normalize_text(column.get("key")) in requested_keys
        ]
        selected_columns[table_key] = filtered_keys or get_default_dashboard_column_keys(table_key)

    return selected_columns


def get_selected_dashboard_column_defs(selected_tables, selected_columns):
    selected_column_defs = OrderedDict()

    for table_key, config in DASHBOARD_TABLE_CONFIGS.items():
        if table_key not in selected_tables:
            continue

        allowed_keys = set(selected_columns.get(table_key) or get_default_dashboard_column_keys(table_key))
        selected_column_defs[table_key] = [
            column
            for column in config.get("columns", [])
            if column.get("key") in allowed_keys
        ]

    return selected_column_defs


def map_sequence_rows(rows, column_keys):
    mapped_rows = []

    for row in rows:
        mapped_rows.append({
            column_key: row[index] if index < len(row) else "-"
            for index, column_key in enumerate(column_keys)
        })

    return mapped_rows


def build_dashboard_table_rows(fish_totals, rice_requirement_rows, fried_rice_rows, combined_menu_rows, detailed_menu_rows):
    return {
        "fish_stock": [
            {
                "no": row["no"],
                "fish_divisi": row["fish_divisi"],
                "fish_name": row["fish_name"],
                "fish_size": row["fish_size"] or "-",
                "fish_weight_display": row["fish_weight_display"] if row["fish_weight"] and row["fish_weight"] > 0 else "-",
                "total": row["total"],
                "menu_names": row["menu_names"] or "-"
            }
            for row in fish_totals
        ],
        "rice_need": map_sequence_rows(rice_requirement_rows, ["no", "keterangan", "jumlah"]),
        "fried_rice": map_sequence_rows(fried_rice_rows, ["no", "menu", "jumlah"]),
        "menu_total": [
            {
                "no": row["no"],
                "divisi": row["divisi"],
                "menu": row["menu"],
                "detail": row["detail"],
                "penyajian_saus": row["penyajian_saus"],
                "special_req": row["special_req"],
                "jumlah": row["jumlah"]
            }
            for row in combined_menu_rows
        ],
        "menu_total_detail": [
            {
                "no": row["no"],
                "divisi": row["divisi"],
                "menu": row["menu"],
                "detail": row["detail"],
                "penyajian_saus": row["penyajian_saus"],
                "special_req": row["special_req"],
                "jumlah": row["jumlah"]
            }
            for row in detailed_menu_rows
        ]
    }


def get_daily_menu_recap_rows(cursor, selected_date, selected_divisi=None, search_query=""):
    description_expr = dish_description_sql("ri")

    query = """
    SELECT
        m.name,
        m.serving_type,
        CASE
            WHEN LOWER(m.name) = 'udang sambal pete' THEN 'lokal'
            WHEN m.divisi IN ('local','bakar/local') THEN 'lokal'
            WHEN m.divisi IN ('seafood','bakar/seafood') THEN 'seafood'
            ELSE m.divisi
        END AS divisi,
        COALESCE(NULLIF(TRIM(ri.fish_type),''),'') AS fish_type,
        ri.fish_size,
        COALESCE(ri.fish_weight,0) AS fish_weight,
        COALESCE(fs.weight_unit, ds.weight_unit, 'ons') AS fish_weight_unit,
    """ + description_expr + """ AS dish_description,
        ri.menu_options_json,
        SUM(ri.quantity) AS total
    FROM reservation_items ri
    JOIN menus m ON m.id = ri.menu_id
    LEFT JOIN fish_stock fs ON fs.id = ri.fish_stock_ref_id
    LEFT JOIN daily_item_stock ds ON ds.id = ri.special_stock_ref_id
    JOIN reservations r ON r.id = ri.reservation_id
    WHERE DATE(r.reservation_datetime) = %s
    """

    params = [selected_date]

    if selected_divisi:
        if selected_divisi == "lokal":
            query += " AND m.divisi IN (%s,%s)"
            params.extend(["local","bakar/local"])
        elif selected_divisi == "seafood":
            query += " AND m.divisi IN (%s,%s)"
            params.extend(["seafood","bakar/seafood"])
        else:
            query += " AND m.divisi = %s"
            params.append(selected_divisi)

    search_query = (search_query or "").strip()
    if search_query:
        query += " AND m.name LIKE %s"
        params.append(f"%{search_query}%")

    query += """
    GROUP BY
        m.name,
        m.serving_type,
        CASE
            WHEN LOWER(m.name) = 'udang sambal pete' THEN 'lokal'
            WHEN m.divisi IN ('local','bakar/local') THEN 'lokal'
            WHEN m.divisi IN ('seafood','bakar/seafood') THEN 'seafood'
            ELSE m.divisi
        END,
        COALESCE(NULLIF(TRIM(ri.fish_type),''),''),
        ri.fish_size,
        COALESCE(ri.fish_weight,0),
        COALESCE(fs.weight_unit, ds.weight_unit, 'ons'),
        ri.menu_options_json,
    """ + description_expr + """
    ORDER BY divisi, m.name, m.serving_type, total DESC
    """

    cursor.execute(query, params)
    rows = cursor.fetchall()

    for index, row in enumerate(rows, start=1):
        row["no"] = index
        row["menu_label"] = get_payload_display_name(
            row.get("menu_options_json"),
            row.get("name"),
            row.get("serving_type")
        )
        row["fish_info"] = format_fish_info(row)
        row["effective_selected_options"] = get_effective_selected_options(row)
        row["option_summary"] = " | ".join(
            f"{(option.get('label') or '').strip()}: {format_option_display_value(option.get('display'))}"
            for option in row["effective_selected_options"].values()
            if (option.get("label") or "").strip() and format_option_display_value(option.get("display"))
        )
        row["display_note"] = " | ".join(
            part for part in [
                row["option_summary"],
                (row.get("dish_description") or "").strip()
            ]
            if part
        )

    return rows


def normalize_base_menu_name(menu_name, serving_type):
    cleaned_name = re.sub(r"\s+", " ", (menu_name or "").strip())
    if not cleaned_name:
        return "-"

    base_name = re.sub(r"\s*\((paket|package|pcs|piece|porsi|portion|menu_piece)\)\s*$", "", cleaned_name, flags=re.IGNORECASE).strip()
    base_name = re.sub(r"\s+(pcs|piece|porsi|portion)\s*$", "", base_name, flags=re.IGNORECASE).strip()

    if normalize_serving_type(serving_type) == "paket":
        base_name = re.sub(r"^(paket|package)\s+", "", base_name, flags=re.IGNORECASE).strip()

    if not base_name:
        return cleaned_name

    return base_name


def get_daily_menu_serving_totals(cursor, selected_date, selected_divisi=None):
    query = """
    SELECT
        m.name,
        m.serving_type,
        SUM(ri.quantity) AS total
    FROM reservation_items ri
    JOIN menus m ON m.id = ri.menu_id
    JOIN reservations r ON r.id = ri.reservation_id
    WHERE DATE(r.reservation_datetime) = %s
    """
    params = [selected_date]

    if selected_divisi:
        if selected_divisi == "lokal":
            query += " AND m.divisi IN (%s,%s)"
            params.extend(["local", "bakar/local"])
        elif selected_divisi == "seafood":
            query += " AND m.divisi IN (%s,%s)"
            params.extend(["seafood", "bakar/seafood"])
        else:
            query += " AND m.divisi = %s"
            params.append(selected_divisi)

    query += """
    GROUP BY m.name, m.serving_type
    ORDER BY m.name, m.serving_type
    """

    cursor.execute(query, params)
    return cursor.fetchall()


def ensure_kitchen_live_tables(cursor):
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS reservation_item_completion (
        id INT AUTO_INCREMENT PRIMARY KEY,
        reservation_item_id INT NOT NULL,
        is_completed TINYINT(1) NOT NULL DEFAULT 0,
        completed_at DATETIME NULL,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
        UNIQUE KEY uniq_reservation_item_completion (reservation_item_id),
        CONSTRAINT fk_reservation_item_completion_item
            FOREIGN KEY (reservation_item_id) REFERENCES reservation_items(id)
            ON DELETE CASCADE
    )
    """)


def get_kitchen_live_reservations(cursor, selected_date, search_query=""):
    ensure_kitchen_live_tables(cursor)
    description_expr = dish_description_sql("ri")
    query = """
    SELECT
        r.id AS reservation_id,
        r.customer_name,
        r.table_number,
        r.people_count,
        r.reservation_datetime,
        r.description AS reservation_description,
        ri.id AS item_id,
        ri.quantity,
        COALESCE(NULLIF(TRIM(ri.fish_type),''),'') AS fish_type,
        ri.fish_size,
        COALESCE(ri.fish_weight,0) AS fish_weight,
        COALESCE(fs.weight_unit, ds.weight_unit, 'ons') AS fish_weight_unit,
    """ + description_expr + """ AS dish_description,
        ri.menu_options_json,
        m.name AS menu_name,
        m.serving_type,
        CASE
            WHEN LOWER(m.name) = 'udang sambal pete' THEN 'lokal'
            WHEN m.divisi IN ('local','bakar/local') THEN 'lokal'
            WHEN m.divisi IN ('seafood','bakar/seafood') THEN 'seafood'
            ELSE m.divisi
        END AS divisi,
        COALESCE(ric.is_completed, 0) AS is_completed
    FROM reservations r
    LEFT JOIN reservation_items ri ON ri.reservation_id = r.id
    LEFT JOIN menus m ON m.id = ri.menu_id
    LEFT JOIN fish_stock fs ON fs.id = ri.fish_stock_ref_id
    LEFT JOIN daily_item_stock ds ON ds.id = ri.special_stock_ref_id
    LEFT JOIN reservation_item_completion ric ON ric.reservation_item_id = ri.id
    WHERE DATE(r.reservation_datetime) = %s
    """
    params = [selected_date]

    if search_query:
        search_like = f"%{search_query}%"
        query += """
        AND (
            r.customer_name LIKE %s
            OR r.table_number LIKE %s
            OR COALESCE(r.description,'') LIKE %s
            OR CAST(r.id AS CHAR) LIKE %s
        )
        """
        params.extend([search_like] * 4)

    query += """
    ORDER BY r.reservation_datetime ASC, r.id ASC, ri.id ASC
    """
    cursor.execute(query, params)

    grouped_reservations = OrderedDict()
    for row in cursor.fetchall():
        reservation_id = row["reservation_id"]
        reservation = grouped_reservations.setdefault(reservation_id, {
            "reservation_id": reservation_id,
            "customer_name": row["customer_name"],
            "table_number": row["table_number"],
            "people_count": row["people_count"],
            "reservation_datetime": row["reservation_datetime"],
            "reservation_description": row.get("reservation_description"),
            "menus": [],
            "total_items": 0,
            "completed_items": 0,
            "all_completed": False
        })

        if row.get("item_id"):
            menu_item = {
                "item_id": row["item_id"],
                "menu_name": row.get("menu_name") or "-",
                "serving_type": row.get("serving_type") or "-",
                "divisi": row.get("divisi") or "-",
                "quantity": row.get("quantity") or 0,
                "dish_description": row.get("dish_description"),
                "menu_display_name": get_payload_display_name(
                    row.get("menu_options_json"),
                    row.get("menu_name"),
                    row.get("serving_type")
                ),
                "option_summary": build_option_summary_from_payload(row.get("menu_options_json")),
                "is_completed": bool(row.get("is_completed"))
            }
            menu_item["fish_info"] = format_fish_info(row)
            reservation["menus"].append(menu_item)

    for reservation in grouped_reservations.values():
        reservation["total_items"] = len(reservation["menus"])
        reservation["completed_items"] = sum(1 for item in reservation["menus"] if item["is_completed"])
        reservation["all_completed"] = reservation["total_items"] > 0 and reservation["completed_items"] == reservation["total_items"]

    return list(grouped_reservations.values())


def to_int_qty(value):
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def build_rice_requirement_rows(menu_totals):
    total_paket = sum(
        to_int_qty(row.get("total"))
        for row in menu_totals
        if normalize_serving_type(row.get("serving_type")) == "paket"
    )
    nasi_putih_total = sum(
        to_int_qty(row.get("total"))
        for row in menu_totals
        if normalize_text(row.get("name")) == "nasi putih"
    )
    total_nasi = total_paket + nasi_putih_total

    return [
        [1, "Total Nasi dari Menu Paket", total_paket],
        [2, "Total Nasi Putih Perpcs", nasi_putih_total],
        [3, "TOTAL KEBUTUHAN NASI", total_nasi]
    ]


def build_fried_rice_rows(menu_totals):
    totals_by_name = {}
    for row in menu_totals:
        name_key = normalize_text(row.get("name"))
        totals_by_name[name_key] = totals_by_name.get(name_key, 0) + to_int_qty(row.get("total"))

    rows = []
    for index, menu_name in enumerate(FRIED_RICE_MENU_NAMES, start=1):
        rows.append([
            index,
            menu_name.title(),
            totals_by_name.get(menu_name, 0)
        ])

    total_fried_rice = sum(item[2] for item in rows)
    rows.append([len(rows) + 1, "Jumlah Sambal Nasgor", total_fried_rice])
    return rows


def increment_counter(counter, value, amount):
    normalized_value = str(value or "").strip()
    if not normalized_value:
        return
    counter[normalized_value] = counter.get(normalized_value, 0) + to_int_qty(amount)


def increment_labeled_counter(counter_map, label, values, amount):
    clean_label = str(label or "").strip()
    if not clean_label:
        return

    bucket = counter_map.setdefault(clean_label, OrderedDict())
    for value in normalize_option_value_list(values):
        increment_counter(bucket, value, amount)


def format_counter_text(counter):
    return format_counter_text_with_style(counter)


def format_counter_text_with_style(counter, use_parentheses=False):
    if not counter:
        return "-"

    ordered_values = sorted(
        counter.items(),
        key=lambda item: (-item[1], normalize_text(item[0]))
    )
    if use_parentheses:
        return ", ".join(f"{value} ({count})" for value, count in ordered_values)
    return ", ".join(f"{value} {count}" for value, count in ordered_values)


def build_option_breakdown_text(option_counter_map):
    return build_option_breakdown_text_with_style(option_counter_map)


def build_option_breakdown_text_with_style(option_counter_map, use_parentheses=False):
    sections = []
    for label, values in option_counter_map.items():
        if not values:
            continue

        value_text = format_counter_text_with_style(values, use_parentheses=use_parentheses)
        sections.append(f"{label}: {value_text}")

    return " | ".join(sections)


def build_menu_dashboard_bucket(menu_name, divisi_name):
    return {
        "menu_name": menu_name,
        "divisi_set": {divisi_name},
        "sambal": OrderedDict(),
        "bumbu": OrderedDict(),
        "topping_saus": OrderedDict(),
        "suhu_minuman": OrderedDict(),
        "pilihan_telur": OrderedDict(),
        "pilihan_saus": OrderedDict(),
        "varian_ukuran": OrderedDict(),
        "opsi_lain": OrderedDict(),
        "size_nila": OrderedDict(),
        "berat_ikan": OrderedDict(),
        "sauce_presentation": OrderedDict(),
        "special_requests": OrderedDict(),
        "total": 0
    }


def apply_row_to_dashboard_bucket(bucket, row):
    qty = to_int_qty(row.get("total"))
    selected_options = row.get("effective_selected_options") or get_effective_selected_options(row)

    for option in selected_options.values():
        label = (option.get("label") or "").strip()
        display_values = normalize_option_value_list(option.get("display"))
        if not label or not display_values:
            continue

        normalized_label = normalize_text(label)
        if normalized_label == normalize_text("Pilih Paket atau PCS"):
            continue

        if normalized_label == normalize_text("Sambal Rica"):
            for display_value in display_values:
                increment_counter(bucket["sambal"], display_value, qty)
            continue

        if normalized_label == normalize_text("Bumbu"):
            for display_value in display_values:
                increment_counter(bucket["bumbu"], display_value, qty)
            continue

        if normalized_label == normalize_text("Topping Saus"):
            for display_value in display_values:
                increment_counter(bucket["topping_saus"], display_value, qty)
            continue

        if normalized_label == normalize_text("Suhu Minuman"):
            for display_value in display_values:
                increment_counter(bucket["suhu_minuman"], display_value, qty)
            continue

        if normalized_label == normalize_text("Pilihan Telur"):
            for display_value in display_values:
                increment_counter(bucket["pilihan_telur"], display_value, qty)
            continue

        if normalized_label == normalize_text("Penyajian Saus"):
            for display_value in display_values:
                increment_counter(bucket["sauce_presentation"], display_value, qty)
            continue

        if normalized_label in (normalize_text("Pilihan Saus"), normalize_text("Penyajian Saus")):
            increment_labeled_counter(bucket["pilihan_saus"], label, display_values, qty)
            continue

        if normalized_label in (normalize_text("Varian Ayam"), normalize_text("Ukuran")):
            increment_labeled_counter(bucket["varian_ukuran"], label, display_values, qty)
            continue

        increment_labeled_counter(bucket["opsi_lain"], label, display_values, qty)

    fish_size = str(row.get("fish_size") or "").strip()
    if fish_size:
        increment_counter(bucket["size_nila"], fish_size.replace("_", " ").title(), qty)

    fish_weight = row.get("fish_weight")
    if fish_weight not in (None, "", 0, 0.0, "0", "0.0"):
        increment_counter(
            bucket["berat_ikan"],
            format_weight_display(fish_weight, row.get("fish_weight_unit")),
            qty
        )

    note = (row.get("dish_description") or "").strip()
    if note and note != "-":
        increment_counter(bucket["special_requests"], note, qty)

    bucket["total"] += qty


def build_dashboard_detail_text(bucket):
    sections = []

    section_map = [
        ("Sambal", bucket["sambal"]),
        ("Bumbu", bucket["bumbu"]),
        ("Topping / Saus", bucket["topping_saus"]),
        ("Suhu", bucket["suhu_minuman"]),
        ("Telur", bucket["pilihan_telur"]),
        ("Size Nila", bucket["size_nila"]),
        ("Berat Ikan", bucket["berat_ikan"])
    ]
    for label, counter in section_map:
        text = format_counter_text_with_style(counter, use_parentheses=True)
        if text != "-":
            sections.append(f"{label}: {text}")

    for option_text in [
        build_option_breakdown_text_with_style(bucket["pilihan_saus"], use_parentheses=True),
        build_option_breakdown_text_with_style(bucket["varian_ukuran"], use_parentheses=True),
        build_option_breakdown_text_with_style(bucket["opsi_lain"], use_parentheses=True)
    ]:
        if option_text:
            sections.append(option_text)

    return " | ".join(sections) or "-"


def build_dashboard_row_output(index, bucket):
    return {
        "no": index,
        "divisi": ", ".join(sorted(bucket["divisi_set"], key=normalize_text)),
        "menu": bucket["menu_name"],
        "detail": build_dashboard_detail_text(bucket),
        "penyajian_saus": format_counter_text_with_style(bucket["sauce_presentation"], use_parentheses=True),
        "special_req": format_counter_text_with_style(bucket["special_requests"], use_parentheses=True),
        "jumlah": bucket["total"]
    }

def build_combined_menu_rows(menu_rows):
    combined_totals = {}

    for row in menu_rows:
        base_name = get_payload_display_name(
            row.get("menu_options_json"),
            row.get("name"),
            row.get("serving_type")
        )
        base_key = normalize_text(base_name)
        divisi_name = (row.get("divisi") or "-").strip() or "-"

        if base_key not in combined_totals:
            combined_totals[base_key] = build_menu_dashboard_bucket(base_name, divisi_name)

        combined_totals[base_key]["divisi_set"].add(divisi_name)
        apply_row_to_dashboard_bucket(combined_totals[base_key], row)

    ordered_rows = sorted(
        combined_totals.values(),
        key=lambda item: (-item["total"], normalize_text(item["menu_name"]))
    )

    return [build_dashboard_row_output(index, row) for index, row in enumerate(ordered_rows, start=1)]


def build_detailed_menu_rows(menu_rows):
    grouped_rows = OrderedDict()

    sorted_rows = sorted(
        menu_rows,
        key=lambda row: (
            normalize_text(row.get("divisi")),
            normalize_text(get_payload_display_name(
                row.get("menu_options_json"),
                row.get("name"),
                row.get("serving_type")
            )),
            normalize_serving_type(row.get("serving_type"))
        )
    )

    for row in sorted_rows:
        base_menu_name = get_payload_display_name(
            row.get("menu_options_json"),
            row.get("name"),
            row.get("serving_type")
        )
        menu_name = format_menu_label(base_menu_name, row.get("serving_type"))
        row_key = normalize_text(menu_name)
        bucket = grouped_rows.setdefault(
            row_key,
            build_menu_dashboard_bucket(menu_name, row.get("divisi") or "-")
        )
        bucket["divisi_set"].add((row.get("divisi") or "-").strip() or "-")
        apply_row_to_dashboard_bucket(bucket, row)

    return [
        build_dashboard_row_output(index, row)
        for index, row in enumerate(grouped_rows.values(), start=1)
    ]


def get_daily_fish_totals(cursor, selected_date):
    cursor.execute("""
    SELECT
        CASE
            WHEN m.divisi IN ('local','bakar/local') THEN 'local'
            WHEN m.divisi IN ('seafood','bakar/seafood') THEN 'seafood'
            ELSE m.divisi
        END AS fish_divisi,
        CASE
            WHEN LOWER(TRIM(COALESCE(ri.fish_type,''))) = 'rahang tuna' THEN 'Rahang Tuna PCS'
            WHEN COALESCE(NULLIF(TRIM(ri.fish_size),''),'') <> '' THEN 'Nila'
            WHEN COALESCE(NULLIF(TRIM(ri.fish_type),''),'') <> '' THEN CONCAT(UCASE(LEFT(ri.fish_type,1)), SUBSTRING(ri.fish_type,2))
            ELSE 'Ikan Laut'
        END AS fish_name,
        NULLIF(TRIM(ri.fish_size),'') AS fish_size,
        COALESCE(ri.fish_weight,0) AS fish_weight,
        COALESCE(fs.weight_unit, ds.weight_unit, 'ons') AS fish_weight_unit,
        m.name,
        m.serving_type,
        SUM(ri.quantity) AS menu_total
    FROM reservation_items ri
    JOIN menus m ON m.id = ri.menu_id
    LEFT JOIN fish_stock fs ON fs.id = ri.fish_stock_ref_id
    LEFT JOIN daily_item_stock ds ON ds.id = ri.special_stock_ref_id
    JOIN reservations r ON r.id = ri.reservation_id
    WHERE DATE(r.reservation_datetime) = %s
    AND (
        m.category IN ('ikan nila','ikan laut')
        OR COALESCE(NULLIF(TRIM(ri.fish_size),''),'') <> ''
        OR COALESCE(NULLIF(TRIM(ri.fish_type),''),'') <> ''
        OR COALESCE(ri.fish_weight,0) > 0
    )
    GROUP BY
        CASE
            WHEN m.divisi IN ('local','bakar/local') THEN 'local'
            WHEN m.divisi IN ('seafood','bakar/seafood') THEN 'seafood'
            ELSE m.divisi
        END,
        CASE
            WHEN LOWER(TRIM(COALESCE(ri.fish_type,''))) = 'rahang tuna' THEN 'Rahang Tuna PCS'
            WHEN COALESCE(NULLIF(TRIM(ri.fish_size),''),'') <> '' THEN 'Nila'
            WHEN COALESCE(NULLIF(TRIM(ri.fish_type),''),'') <> '' THEN CONCAT(UCASE(LEFT(ri.fish_type,1)), SUBSTRING(ri.fish_type,2))
            ELSE 'Ikan Laut'
        END,
        NULLIF(TRIM(ri.fish_size),''),
        COALESCE(ri.fish_weight,0),
        COALESCE(fs.weight_unit, ds.weight_unit, 'ons'),
        m.name,
        m.serving_type
    ORDER BY fish_divisi, fish_name, fish_size, fish_weight, m.name, m.serving_type
    """,(selected_date,))

    grouped_rows = OrderedDict()
    for row in cursor.fetchall():
        group_key = (
            row["fish_divisi"],
            row["fish_name"],
            row["fish_size"],
            row["fish_weight"],
            row["fish_weight_unit"]
        )

        if group_key not in grouped_rows:
            grouped_rows[group_key] = {
                "fish_divisi": row["fish_divisi"],
                "fish_name": row["fish_name"],
                "fish_size": row["fish_size"],
                "fish_weight": row["fish_weight"],
                "fish_weight_unit": row["fish_weight_unit"],
                "total": 0,
                "menu_totals": OrderedDict()
            }

        menu_label = format_menu_label(row.get("name"), row.get("serving_type"))
        menu_total = to_int_qty(row.get("menu_total"))
        grouped_rows[group_key]["total"] += menu_total
        grouped_rows[group_key]["menu_totals"][menu_label] = (
            grouped_rows[group_key]["menu_totals"].get(menu_label, 0) + menu_total
        )

    rows = list(grouped_rows.values())
    for row in rows:
        row["fish_weight_display"] = format_weight_display(row["fish_weight"], row.get("fish_weight_unit"))
        row["menu_names"] = "\n".join(
            f"{menu_name} ({menu_total})"
            for menu_name, menu_total in row["menu_totals"].items()
        ) or "-"
    return rows


def status_indicator(status):
    if status == "ready":
        return "green"
    if status == "pending":
        return "yellow"
    return "red"


def effective_stock_status(status, qty=None):
    normalized_status = (status or "").strip().lower() or "not_ready"

    if normalized_status != "ready":
        return normalized_status

    if qty is not None and int(qty or 0) <= 0:
        return "not_ready"

    return "ready"


def normalize_weight_to_ons(weight_value, weight_unit="ons"):
    if weight_value in (None, ""):
        return 0

    weight = float(weight_value)
    if weight_unit == "kg":
        return round(weight * 10, 1)
    return round(weight, 1)


def format_weight_display(weight_ons, weight_unit="ons"):
    if not weight_ons or float(weight_ons) <= 0:
        return "-"

    weight_ons = float(weight_ons)
    if weight_unit == "kg":
        display_value = round(weight_ons / 10, 1)
        return f"{display_value:g} kg"
    return f"{weight_ons:g} ons"


def render_notice_page(title, message, back_url=None, back_label="Kembali", status_code=404):
    fallback_url = back_url or request.referrer or url_for("reservations")
    return render_template(
        "notice.html",
        title=title,
        message=message,
        back_url=fallback_url,
        back_label=back_label
    ), status_code


def row_value(row, key, index=0, default=None):
    if row is None:
        return default
    if isinstance(row, dict):
        return row.get(key, default)
    if isinstance(row, (list, tuple)) and len(row) > index:
        return row[index]
    return default


def apply_special_stock_statuses(menus, piece_stock_map, special_tuna_stock):
    package_stock = special_tuna_stock.get("package_stock") or {}
    package_status = effective_stock_status(
        package_stock.get("status"),
        package_stock.get("available_qty")
    )

    rahang_summary = special_tuna_stock.get("summary") or {}
    rahang_status = effective_stock_status(
        rahang_summary.get("status"),
        rahang_summary.get("available_qty")
    )

    for menu in menus:
        menu_name = (menu.get("name") or "").strip().lower()
        stock_source = menu.get("stock_source")

        if (stock_source == "menu_piece" and menu.get("id") in piece_stock_map) or menu_name in PACKAGE_TUNA_MENU_NAMES:
            menu["status"] = package_status
        elif stock_source == "tuna_weight" or menu_name == "rahang tuna":
            menu["status"] = rahang_status


def get_stock_context(cursor, selected_date):
    ensure_additional_stock_tables(cursor)
    ensure_menu_catalog_updates(cursor)
    latest_menu_status_map = get_latest_menu_status_map(cursor, selected_date)

    cursor.execute("""
    SELECT
        m.id,
        m.name,
        m.serving_type,
        m.category,
        m.divisi,
        m.stock_type,
        CASE
            WHEN LOWER(m.name) = 'rahang tuna' THEN 'tuna_weight'
            WHEN LOWER(m.name) IN ('dada tuna goreng','dada tuna bakar','paket dada tuna goreng','paket dada tuna bakar') THEN 'menu_piece'
            ELSE m.stock_type
        END AS stock_source,
        'ready' AS status
    FROM menus m
    ORDER BY m.name
    """)
    menus = cursor.fetchall()

    for menu in menus:
        latest_status_row = latest_menu_status_map.get(menu["id"])
        if latest_status_row:
            menu["status"] = row_value(latest_status_row, "status", 1, "ready") or "ready"

    latest_nila_status_map = get_latest_nila_status_map(cursor, selected_date)
    nila_sizes = [
        {"size_category": size_key}
        for size_key in ["kecil","sedang","besar","jumbo","super_jumbo"]
        if effective_stock_status(row_value(latest_nila_status_map.get(size_key), "status", 2, "ready")) == "ready"
    ]

    sea_fish = [
        row
        for row in get_latest_sea_fish_stock_rows(cursor, selected_date)
        if row.get("status") == "ready" and int(row.get("fish_count") or 0) > 0
    ]

    _, piece_stock_map, special_tuna_stock = get_special_tuna_stock_context(cursor, selected_date)
    apply_special_stock_statuses(menus, piece_stock_map, special_tuna_stock)

    option_stock_map = get_latest_option_stock_map(cursor, selected_date)
    display_catalog = build_display_menu_catalog(menus)
    apply_option_stock_to_catalog(display_catalog, option_stock_map)

    return display_catalog, nila_sizes, sea_fish


def ensure_additional_stock_tables(cursor):
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS daily_item_stock (
        id INT AUTO_INCREMENT PRIMARY KEY,
        menu_id INT NOT NULL,
        stock_date DATE NOT NULL,
        weight_ons DECIMAL(5,1) NOT NULL DEFAULT 0.0,
        weight_unit ENUM('ons','kg') DEFAULT 'ons',
        available_qty INT NOT NULL DEFAULT 0,
        status ENUM('ready','not_ready') DEFAULT 'ready',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
        UNIQUE KEY uniq_menu_stock (menu_id, stock_date, weight_ons),
        CONSTRAINT fk_daily_item_stock_menu FOREIGN KEY (menu_id) REFERENCES menus(id) ON DELETE CASCADE
    )
    """)

    cursor.execute("""
    ALTER TABLE fish_stock
    ADD COLUMN IF NOT EXISTS weight_unit ENUM('ons','kg') DEFAULT 'ons'
    """)

    cursor.execute("""
    ALTER TABLE daily_item_stock
    ADD COLUMN IF NOT EXISTS weight_unit ENUM('ons','kg') DEFAULT 'ons'
    """)

    cursor.execute("SHOW COLUMNS FROM daily_menu_stock LIKE 'status'")
    daily_menu_status_column = cursor.fetchone()
    daily_menu_status_type = (row_value(daily_menu_status_column, "Type", 1, "") or "").lower()
    if daily_menu_status_type and ("not_ready" not in daily_menu_status_type or "'out'" in daily_menu_status_type):
        cursor.execute("""
        ALTER TABLE daily_menu_stock
        MODIFY COLUMN status ENUM('ready','pending','out','not_ready') DEFAULT 'ready'
        """)
        cursor.execute("""
        UPDATE daily_menu_stock
        SET status = 'not_ready'
        WHERE status IN ('out','')
        """)
        cursor.execute("""
        ALTER TABLE daily_menu_stock
        MODIFY COLUMN status ENUM('ready','pending','not_ready') DEFAULT 'ready'
        """)

    cursor.execute("""
    ALTER TABLE reservation_items
    ADD COLUMN IF NOT EXISTS fish_stock_ref_id INT NULL
    """)

    cursor.execute("""
    ALTER TABLE reservation_items
    ADD COLUMN IF NOT EXISTS special_stock_ref_id INT NULL
    """)

    cursor.execute("""
    ALTER TABLE reservation_items
    ADD COLUMN IF NOT EXISTS menu_options_json LONGTEXT NULL
    """)

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS menu_option_stock (
        id INT AUTO_INCREMENT PRIMARY KEY,
        option_key VARCHAR(80) NOT NULL,
        option_value VARCHAR(80) NOT NULL,
        status ENUM('ready','not_ready') NOT NULL DEFAULT 'ready',
        stock_date DATE NOT NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
        UNIQUE KEY uniq_option_stock (option_key, option_value, stock_date)
    )
    """)

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS stock_change_log (
        id INT AUTO_INCREMENT PRIMARY KEY,
        stock_scope VARCHAR(80) NOT NULL,
        target_name VARCHAR(255) NOT NULL,
        previous_value TEXT NULL,
        new_value TEXT NULL,
        actor_name VARCHAR(120) NULL,
        notes TEXT NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """)

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS reservation_change_log (
        id INT AUTO_INCREMENT PRIMARY KEY,
        reservation_id INT NULL,
        reservation_item_id INT NULL,
        action_type VARCHAR(40) NOT NULL,
        change_scope VARCHAR(40) NOT NULL,
        summary TEXT NOT NULL,
        actor_name VARCHAR(120) NULL,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """)

    ensure_menu_catalog_updates(cursor)


def get_special_tuna_stock_context(cursor, selected_date):
    ensure_additional_stock_tables(cursor)

    cursor.execute("""
    SELECT
        id,
        name,
        serving_type
    FROM menus
    WHERE LOWER(name) IN (
        'rahang tuna',
        'dada tuna goreng',
        'dada tuna bakar',
        'paket dada tuna goreng',
        'paket dada tuna bakar'
    )
    ORDER BY
        CASE
            WHEN LOWER(name) IN ('paket dada tuna goreng','paket dada tuna bakar') THEN 1
            WHEN LOWER(name) IN ('dada tuna goreng','dada tuna bakar') THEN 2
            WHEN LOWER(name) = 'rahang tuna' THEN 3
            ELSE 4
        END,
        name
    """)
    stock_menus = cursor.fetchall()

    if not stock_menus:
        return [], {}, {}

    menu_ids = [row["id"] for row in stock_menus]
    stock_rows = get_latest_daily_item_stock_rows(cursor, menu_ids, selected_date)

    package_menu = None
    rahang_menu = None
    package_rows = []
    ready_rahang_rows = []
    all_rahang_rows = []
    piece_stock_map = {}

    for stock_menu in stock_menus:
        name_key = stock_menu["name"].strip().lower()
        if name_key == "rahang tuna":
            rahang_menu = stock_menu
        elif package_menu is None and name_key in PACKAGE_TUNA_MENU_NAMES:
            package_menu = stock_menu

    for row in stock_rows:
        name_key = row["name"].strip().lower()
        row["status"] = effective_stock_status(row.get("status"), row.get("available_qty"))

        if name_key == "rahang tuna":
            row["display_weight"] = format_weight_display(row["weight_ons"], row.get("weight_unit"))
            row["label"] = f"Rahang Tuna - {row['display_weight']} ({row['available_qty']} stok)"
            all_rahang_rows.append(row)
            if row["weight_ons"] and row["available_qty"] > 0 and row["status"] == "ready":
                ready_rahang_rows.append(row)
        else:
            package_rows.append(row)

    package_stock = None
    if package_menu:
        package_stock = next((row for row in package_rows if row["menu_id"] == package_menu["id"]), None)
    if package_stock is None and package_rows:
        package_stock = package_rows[0]

    if package_stock:
        package_stock["display_name"] = "Paket Dada Tuna"
        package_stock["status_dot"] = status_indicator(package_stock["status"])

    for row in stock_rows:
        name_key = row["name"].strip().lower()

        if name_key != "rahang tuna" and package_stock:
            piece_stock_map[row["menu_id"]] = package_stock

    if package_stock:
        for stock_menu in stock_menus:
            if stock_menu["name"].strip().lower() in PACKAGE_TUNA_MENU_NAMES:
                piece_stock_map[stock_menu["id"]] = package_stock

    ready_rahang_total = sum(row["available_qty"] for row in ready_rahang_rows)
    rahang_status_row = next(
        (row for row in all_rahang_rows if (row["weight_ons"] or 0) == 0 and row["status"] == "not_ready"),
        None
    )
    rahang_summary = {
        "display_name": "Rahang Tuna PCS",
        "available_qty": ready_rahang_total,
        "status": "ready" if ready_rahang_total > 0 else "not_ready",
        "status_dot": status_indicator("ready" if ready_rahang_total > 0 else "not_ready"),
        "source_count": len(ready_rahang_rows),
        "has_marker": bool(rahang_status_row)
    }
    if ready_rahang_total <= 0 and rahang_status_row:
        rahang_summary["status"] = "not_ready"
        rahang_summary["status_dot"] = status_indicator("not_ready")

    filtered_stock_menus = [package_menu] if package_menu else []

    return filtered_stock_menus, piece_stock_map, {
        "menu": rahang_menu,
        "rows": ready_rahang_rows,
        "all_rows": all_rahang_rows,
        "summary": rahang_summary,
        "package_stock": package_stock
    }


def reduce_stock_after_order(cursor, fish_stock_id=None, special_stock_id=None, qty=0):
    if qty <= 0:
        return

    if fish_stock_id:
        cursor.execute("""
        UPDATE fish_stock
        SET fish_count = GREATEST(fish_count - %s, 0)
        WHERE id = %s
        """,(qty, fish_stock_id))

    if special_stock_id:
        cursor.execute("""
        UPDATE daily_item_stock
        SET available_qty = GREATEST(available_qty - %s, 0)
        WHERE id = %s
        """,(qty, special_stock_id))


def resolve_package_stock_row(cursor, stock_date):
    placeholders = ",".join(["%s"] * len(PACKAGE_TUNA_MENU_NAMES))
    cursor.execute(
        f"""
        SELECT
            ds.id,
            ds.menu_id,
            ds.available_qty,
            ds.status
        FROM daily_item_stock ds
        JOIN menus m ON m.id = ds.menu_id
        WHERE ds.stock_date <= %s
        AND LOWER(m.name) IN ({placeholders})
        AND COALESCE(ds.weight_ons, 0) = 0
        ORDER BY
            ds.stock_date DESC,
            CASE WHEN ds.status = 'ready' AND ds.available_qty > 0 THEN 0 ELSE 1 END,
            ds.id
        LIMIT 1
        """,
        [stock_date] + list(PACKAGE_TUNA_MENU_NAMES)
    )
    row = cursor.fetchone()
    if row and effective_stock_status(row_value(row, "status", 3), row_value(row, "available_qty", 2)) == "ready":
        return row
    return None


def resolve_selected_stock_refs(cursor, reservation_date, menu_id, fish_type=None, fish_weight=None, fish_stock_id=None, special_stock_id=None):
    resolved_fish_stock_id = int(fish_stock_id) if fish_stock_id else None
    resolved_special_stock_id = int(special_stock_id) if special_stock_id else None

    if not reservation_date:
        return resolved_fish_stock_id, resolved_special_stock_id

    normalized_fish_type = (fish_type or "").strip().lower()
    normalized_weight = fish_weight

    if not resolved_special_stock_id:
        cursor.execute("SELECT LOWER(name) AS name FROM menus WHERE id = %s",(menu_id,))
        selected_menu = cursor.fetchone()
        menu_name = str(row_value(selected_menu, "name", 0, "")).strip().lower()

        if menu_name in PACKAGE_TUNA_MENU_NAMES:
            package_row = resolve_package_stock_row(cursor, reservation_date)
            if package_row:
                resolved_special_stock_id = row_value(package_row, "id", 0)

    if not resolved_special_stock_id and normalized_fish_type == "rahang tuna" and normalized_weight not in (None, "", "0", "0.0"):
        cursor.execute("""
        SELECT ds.id
        FROM daily_item_stock ds
        JOIN menus m ON m.id = ds.menu_id
        WHERE ds.stock_date <= %s
        AND LOWER(m.name) = 'rahang tuna'
        AND ds.weight_ons = %s
        ORDER BY ds.stock_date DESC, ds.id DESC
        LIMIT 1
        """,(reservation_date, normalized_weight))
        row = cursor.fetchone()
        if row:
            resolved_special_stock_id = row_value(row, "id", 0)

    if not resolved_fish_stock_id and normalized_fish_type and normalized_fish_type != "rahang tuna" and normalized_weight not in (None, "", "0", "0.0"):
        cursor.execute("""
        SELECT fs.id
        FROM fish_stock fs
        JOIN fish_types ft ON ft.id = fs.fish_type_id
        WHERE fs.stock_date <= %s
        AND LOWER(ft.name) = %s
        AND fs.weight_ons = %s
        ORDER BY fs.stock_date DESC, fs.id DESC
        LIMIT 1
        """,(reservation_date, normalized_fish_type, normalized_weight))
        row = cursor.fetchone()
        if row:
            resolved_fish_stock_id = row_value(row, "id", 0)

    return resolved_fish_stock_id, resolved_special_stock_id


def resolve_item_stock_refs(cursor, item):
    fish_stock_id = item.get("fish_stock_ref_id")
    special_stock_id = item.get("special_stock_ref_id")

    if fish_stock_id or special_stock_id:
        return fish_stock_id, special_stock_id

    reservation_date = item.get("reservation_date")
    fish_type = (item.get("fish_type") or "").strip().lower()
    fish_weight = item.get("fish_weight") or 0
    menu_name = (item.get("name") or "").strip().lower()

    if not reservation_date:
        return None, None

    if fish_type and fish_weight and fish_type != "none":
        if fish_type == "rahang tuna":
            cursor.execute("""
            SELECT ds.id
            FROM daily_item_stock ds
            JOIN menus m ON m.id = ds.menu_id
            WHERE ds.stock_date <= %s
            AND LOWER(m.name) = 'rahang tuna'
            AND ds.weight_ons = %s
            ORDER BY ds.stock_date DESC, ds.id DESC
            LIMIT 1
            """,(reservation_date, fish_weight))
            row = cursor.fetchone()
            return None, row_value(row, "id", 0) if row else None

        cursor.execute("""
        SELECT fs.id
        FROM fish_stock fs
        JOIN fish_types ft ON ft.id = fs.fish_type_id
        WHERE fs.stock_date <= %s
        AND LOWER(ft.name) = %s
        AND fs.weight_ons = %s
        ORDER BY fs.stock_date DESC, fs.id DESC
        LIMIT 1
        """,(reservation_date, fish_type, fish_weight))
        row = cursor.fetchone()
        return row_value(row, "id", 0) if row else None, None

    if menu_name in PACKAGE_TUNA_MENU_NAMES:
        placeholders = ",".join(["%s"] * len(PACKAGE_TUNA_MENU_NAMES))
        cursor.execute(
            f"""
            SELECT ds.id
            FROM daily_item_stock ds
            JOIN menus m ON m.id = ds.menu_id
            WHERE ds.stock_date <= %s
            AND LOWER(m.name) IN ({placeholders})
            AND COALESCE(ds.weight_ons, 0) = 0
            ORDER BY ds.stock_date DESC, ds.id DESC
            LIMIT 1
            """,
            [reservation_date] + list(PACKAGE_TUNA_MENU_NAMES)
        )
        row = cursor.fetchone()
        return None, row_value(row, "id", 0) if row else None

    return None, None


def validate_stock_request(cursor, reservation_date, menu_id, qty, fish_type=None, fish_size=None, fish_weight=None, fish_stock_id=None, special_stock_id=None, current_item=None, menu_options_json=None):
    try:
        qty = int(qty)
    except (TypeError, ValueError):
        return False, "Jumlah menu tidak valid."

    if qty <= 0:
        return False, "Jumlah menu harus lebih dari 0."

    is_option_ready, option_message = validate_menu_option_stock(cursor, reservation_date, menu_options_json)
    if not is_option_ready:
        return False, option_message

    cursor.execute("""
    SELECT
        LOWER(name) AS name,
        CASE
            WHEN LOWER(name) = 'rahang tuna' THEN 'tuna_weight'
            WHEN LOWER(name) IN ('dada tuna goreng','dada tuna bakar','paket dada tuna goreng','paket dada tuna bakar') THEN 'menu_piece'
            ELSE stock_type
        END AS stock_source
    FROM menus
    WHERE id = %s
    """,(menu_id,))
    menu_row = cursor.fetchone()

    if not menu_row:
        return False, "Menu tidak ditemukan."

    stock_source = row_value(menu_row, "stock_source", 1)
    menu_name = str(row_value(menu_row, "name", 0, "")).strip().lower()
    current_qty = int(current_item.get("quantity") or 0) if current_item else 0
    current_fish_ref = current_item.get("fish_stock_ref_id") if current_item else None
    current_special_ref = current_item.get("special_stock_ref_id") if current_item else None

    if stock_source == "size":
        if not fish_size:
            return False, "Ukuran Nila belum dipilih."

        nila_row = get_latest_nila_status_map(cursor, reservation_date).get(fish_size)
        nila_status = row_value(nila_row, "status", 0, "not_ready")
        if effective_stock_status(nila_status) != "ready":
            return False, f"Stock Nila ukuran {fish_size} sedang tidak ready."
        return True, None

    if stock_source == "weight":
        if not fish_stock_id:
            return False, "Stock ikan laut belum dipilih."

        cursor.execute("""
        SELECT fish_count, status, stock_date
        FROM fish_stock
        WHERE id = %s
        LIMIT 1
        """,(fish_stock_id,))
        stock_row = cursor.fetchone()
        if not stock_row:
            return False, "Stock ikan laut tidak ditemukan."

        available = int(row_value(stock_row, "fish_count", 0, 0) or 0)
        status = row_value(stock_row, "status", 1, "not_ready")

        if current_fish_ref and str(current_fish_ref) == str(fish_stock_id):
            available += current_qty

        if effective_stock_status(status, available) != "ready":
            return False, "Stock ikan laut sedang tidak ready."

        if qty > available:
            return False, f"Stock ikan laut hanya tersisa {available}."

        return True, None

    if stock_source in ("tuna_weight", "menu_piece"):
        if not special_stock_id:
            stock_label = "Rahang Tuna" if stock_source == "tuna_weight" or menu_name == "rahang tuna" else "Paket Dada Tuna"
            return False, f"Stock {stock_label} belum dipilih."

        cursor.execute("""
        SELECT available_qty, status, stock_date
        FROM daily_item_stock
        WHERE id = %s
        LIMIT 1
        """,(special_stock_id,))
        stock_row = cursor.fetchone()
        if not stock_row:
            return False, "Stock item khusus tidak ditemukan."

        available = int(row_value(stock_row, "available_qty", 0, 0) or 0)
        status = row_value(stock_row, "status", 1, "not_ready")

        if current_special_ref and str(current_special_ref) == str(special_stock_id):
            available += current_qty

        if effective_stock_status(status, available) != "ready":
            label = "Rahang Tuna" if stock_source == "tuna_weight" or menu_name == "rahang tuna" else "Paket Dada Tuna"
            return False, f"Stock {label} sedang tidak ready."

        if qty > available:
            label = "Rahang Tuna" if stock_source == "tuna_weight" or menu_name == "rahang tuna" else "Paket Dada Tuna"
            return False, f"Stock {label} hanya tersisa {available}."

    return True, None


def restore_stock_for_item(cursor, item):
    qty = int(item.get("quantity") or 0)
    if qty <= 0:
        return

    fish_stock_id = item.get("fish_stock_ref_id")
    special_stock_id = item.get("special_stock_ref_id")

    # Hindari restore dobel dari data lama yang belum pernah menyimpan ref stok.
    if not fish_stock_id and not special_stock_id:
        return

    if fish_stock_id:
        cursor.execute("""
        UPDATE fish_stock
        SET fish_count = fish_count + %s
        WHERE id = %s
        """,(qty, fish_stock_id))

    if special_stock_id:
        cursor.execute("""
        UPDATE daily_item_stock
        SET available_qty = available_qty + %s
        WHERE id = %s
        """,(qty, special_stock_id))


# ================= LOGIN =================

@app.route("/restaurant")
def restaurant_landing_legacy():
    return redirect(url_for("restaurant_landing"), code=301)


@app.route("/booking-in-manna")
def restaurant_landing():
    next_slot = (datetime.now() + timedelta(days=1)).replace(minute=0, second=0, microsecond=0)

    return render_template(
        "restaurant/landing.html",
        restaurant=RESTAURANT_PROFILE,
        whatsapp_link=get_restaurant_whatsapp_link(),
        suggested_datetime=next_slot.strftime("%Y-%m-%dT%H:%M"),
        suggested_min_datetime=datetime.now().strftime("%Y-%m-%dT%H:%M")
    )


@app.route("/restaurant/book", methods=["POST"])
def restaurant_book_legacy():
    return redirect(url_for("restaurant_landing") + "#booking", code=307)


@app.route("/booking-in-manna/book", methods=["POST"])
def restaurant_book():
    customer_name = (request.form.get("customer_name") or "").strip()
    whatsapp_number = (request.form.get("whatsapp_number") or "").strip()
    people_count_raw = (request.form.get("people_count") or "").strip()
    reservation_datetime = (request.form.get("reservation_datetime") or "").strip()
    seating_preference = (request.form.get("seating_preference") or "").strip()
    notes = (request.form.get("notes") or "").strip()

    if not customer_name or not whatsapp_number or not people_count_raw or not reservation_datetime:
        flash("Mohon lengkapi nama, WhatsApp, jumlah tamu, dan waktu reservasi.", "error")
        return redirect(url_for("restaurant_landing") + "#booking")

    try:
        people_count = int(people_count_raw)
        if people_count <= 0:
            raise ValueError
    except ValueError:
        flash("Jumlah tamu harus berupa angka lebih dari 0.", "error")
        return redirect(url_for("restaurant_landing") + "#booking")

    try:
        parsed_reservation_datetime = datetime.strptime(reservation_datetime, "%Y-%m-%dT%H:%M")
    except ValueError:
        flash("Format tanggal dan jam reservasi tidak valid.", "error")
        return redirect(url_for("restaurant_landing") + "#booking")

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            INSERT INTO reservations
            (customer_name, table_number, people_count, reservation_datetime, description)
            VALUES (%s,%s,%s,%s,%s)
        """,(
            customer_name,
            (seating_preference[:20] if seating_preference else "BOOKING-WEB"),
            people_count,
            parsed_reservation_datetime.strftime("%Y-%m-%d %H:%M:%S"),
            build_public_booking_description(whatsapp_number, seating_preference, notes)
        ))

        reservation_id = cursor.lastrowid
        conn.commit()
    except mysql.connector.Error:
        conn.rollback()
        flash("Booking belum bisa diproses. Silakan coba lagi beberapa saat.", "error")
        return redirect(url_for("restaurant_landing") + "#booking")
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for("restaurant_booking_success", reservation_id=reservation_id))


@app.route("/restaurant/booking-success/<int:reservation_id>")
def restaurant_booking_success_legacy(reservation_id):
    return redirect(url_for("restaurant_booking_success", reservation_id=reservation_id), code=301)


@app.route("/booking-in-manna/booking-success/<int:reservation_id>")
def restaurant_booking_success(reservation_id):
    return render_template(
        "restaurant/booking_success.html",
        restaurant=RESTAURANT_PROFILE,
        reservation_id=reservation_id,
        whatsapp_link=get_restaurant_whatsapp_link(
            f"Halo Manna Bakery and Cafe, saya ingin konfirmasi booking dengan kode #{reservation_id}."
        )
    )


@app.route("/login", methods=["GET","POST"])
def login():

    if request.method == "POST":

        username = request.form["username"]
        password = request.form["password"]

        user = users.get(username)

        if user and user["password"] == password:
            login_user(User(username, user["role"]))
            if user["role"] == "admin":
                return redirect(url_for("restaurant_landing"))
            return redirect(url_for("home"))

        flash("Login gagal")

    return render_template("login.html")


@app.errorhandler(404)
def handle_not_found(error):
    return render_notice_page(
        "Halaman Tidak Ditemukan",
        "Halaman yang Anda cari tidak tersedia atau sudah dipindahkan.",
        status_code=404
    )


@app.errorhandler(500)
def handle_server_error(error):
    return render_notice_page(
        "Terjadi Kesalahan",
        "Sistem sedang mengalami kendala. Silakan kembali ke halaman sebelumnya lalu coba lagi.",
        status_code=500
    )


# ================= LOGOUT =================

@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


# ================= HOME =================

@app.route("/")
@login_required
def home():
    selected_date = request.args.get("date")

    if not selected_date:
        selected_date = datetime.now().strftime("%Y-%m-%d")

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    _, _, special_tuna_stock = get_special_tuna_stock_context(cursor, selected_date)
    latest_menu_status_map = get_latest_menu_status_map(cursor, selected_date)
    latest_nila_status_map = get_latest_nila_status_map(cursor, selected_date)

    cursor.execute("""
        SELECT COUNT(*) AS total_res
        FROM reservations
        WHERE DATE(reservation_datetime)=%s
    """,(selected_date,))
    total_res = cursor.fetchone()["total_res"]

    cursor.execute("""
        SELECT SUM(people_count) AS total_pax
        FROM reservations
        WHERE DATE(reservation_datetime)=%s
    """,(selected_date,))
    total_pax = cursor.fetchone()["total_pax"] or 0

    busy = total_pax > 80

    cursor.execute("SELECT id, name FROM menus ORDER BY name")
    stock_today = []
    for menu_row in cursor.fetchall():
        latest_status_row = latest_menu_status_map.get(menu_row["id"])
        current_status = row_value(latest_status_row, "status", 1, "ready")
        if current_status in ("not_ready", "pending", "out"):
            stock_today.append({
                "name": menu_row["name"],
                "status": current_status
            })
    stock_today = [
        row for row in stock_today
        if (row.get("name") or "").strip().lower() not in SPECIAL_TUNA_MENU_NAMES
    ]

    sea_fish_stock = [
        row
        for row in get_latest_sea_fish_stock_rows(cursor, selected_date)
        if float(row.get("weight_ons") or 0) > 0 or int(row.get("fish_count") or 0) > 0
    ]

    for fish in sea_fish_stock:
        fish["display_weight"] = format_weight_display(fish["weight_ons"], fish.get("weight_unit"))

    nila_stock = []

    for size in ["kecil","sedang","besar","jumbo","super_jumbo"]:
        current_status = row_value(latest_nila_status_map.get(size), "status", 2, "ready")
        nila_stock.append({
            "size_category": size,
            "status": current_status,
            "status_dot": status_indicator(current_status)
        })

    special_stock = []
    rahang_tuna_rows = []
    package_stock = special_tuna_stock.get("package_stock")
    if package_stock:
        special_stock.append({
            "name": "Paket Dada Tuna",
            "available_qty": package_stock.get("available_qty", 0),
            "status": package_stock.get("status", "ready"),
            "status_dot": status_indicator(package_stock.get("status", "ready"))
        })

    for row in special_tuna_stock.get("all_rows", []):
        if row.get("weight_ons") and float(row["weight_ons"]) > 0:
            rahang_tuna_rows.append({
                "display_weight": format_weight_display(row["weight_ons"], row.get("weight_unit")),
                "available_qty": row.get("available_qty", 0),
                "status": row.get("status", "ready")
            })

    package_status = effective_stock_status(
        package_stock.get("status") if package_stock else "not_ready",
        package_stock.get("available_qty") if package_stock else 0
    )
    if package_status != "ready":
        stock_today.append({
            "name": "Paket Dada Tuna",
            "status": package_status
        })

    rahang_summary = special_tuna_stock.get("summary") or {}
    if effective_stock_status(rahang_summary.get("status"), rahang_summary.get("available_qty")) != "ready":
        stock_today.append({
            "name": "Rahang Tuna",
            "status": effective_stock_status(rahang_summary.get("status"), rahang_summary.get("available_qty"))
        })

    cursor.close()
    conn.close()

    return render_template(
        "home.html",
        total_reservations=total_res,
        total_pax=total_pax,
        busy=busy,
        selected_date=selected_date,
        stock_today=stock_today,
        sea_fish_stock=sea_fish_stock,
        nila_stock=nila_stock,
        special_stock=special_stock,
        rahang_tuna_rows=rahang_tuna_rows,
        rahang_tuna_summary=special_tuna_stock.get("summary", {})
    )

# ================= CREATE RESERVATION PAGE =================

@app.route("/create_reservation")
@login_required
def create_reservation():

    selected_date = request.args.get("date")

    if not selected_date:
        selected_date = date.today().strftime("%Y-%m-%d")

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    menus, nila_sizes, sea_fish = get_stock_context(cursor, selected_date)
    tuna_stock_menus, tuna_piece_stock, rahang_tuna_stock = get_special_tuna_stock_context(cursor, selected_date)

    nila_size_options = [row["size_category"] for row in nila_sizes]

    for fish in sea_fish:
        fish["display_weight"] = format_weight_display(fish["weight_ons"], fish.get("weight_unit"))
        fish["label"] = f"{fish['name'].capitalize()} - {fish['display_weight']} ({fish['fish_count']} stok)"

    cursor.close()
    conn.close()

    return render_template(
        "create_reservation.html",
        menus=menus,
        nila_sizes=nila_size_options,
        sea_fish=sea_fish,
        tuna_piece_stock=tuna_piece_stock,
        rahang_tuna_stock=rahang_tuna_stock,
        selected_date=selected_date
    )


# ================= ADD RESERVATION =================

# ================= ADD RESERVATION =================

@app.route("/add_reservation", methods=["POST"])
@login_required
def add_reservation():

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    ensure_additional_stock_tables(cursor)

    name = request.form["customer_name"]
    table = request.form["table_number"]
    people = request.form["people_count"]
    datetime_res = request.form["reservation_datetime"]
    description = request.form["description"]
    reservation_date = datetime_res.split("T")[0] if "T" in datetime_res else datetime_res.split(" ")[0]
    

    cursor.execute("""
        INSERT INTO reservations
        (customer_name, table_number, people_count, reservation_datetime, description)
        VALUES (%s,%s,%s,%s,%s)
    """,(name,table,people,datetime_res,description))

    reservation_id = cursor.lastrowid
    log_reservation_history(
        cursor,
        reservation_id,
        "create",
        "reservation",
        f"Reservasi dibuat untuk {name}, meja {table}, pax {people}.",
        actor=getattr(current_user, "id", "system")
    )

    menu_ids = request.form.getlist("menu_id[]")
    display_menu_ids = request.form.getlist("display_menu_id[]")
    fish_types = request.form.getlist("fish_type[]")
    fish_sizes = request.form.getlist("fish_size[]")
    fish_weights = request.form.getlist("fish_weight[]")
    fish_stock_ids = request.form.getlist("fish_stock_id[]")
    special_stock_ids = request.form.getlist("special_stock_id[]")
    qtys = request.form.getlist("qty[]")
    special_requests = request.form.getlist("special_request[]")
    dish_descriptions = request.form.getlist("dish_description[]")
    menu_options_json_list = request.form.getlist("menu_options_json[]")
    menu_catalog, _, _ = get_stock_context(cursor, reservation_date)

    for menu_id, display_menu_id, qty, special, dish_desc, menu_options_json, fish_type, size, weight, fish_stock_id, special_stock_id in zip(
menu_ids, display_menu_ids, qtys, special_requests, dish_descriptions, menu_options_json_list, fish_types, fish_sizes, fish_weights, fish_stock_ids, special_stock_ids):

        resolved_menu_id = resolve_menu_submission(
            menu_catalog,
            menu_id,
            display_menu_id=display_menu_id,
            menu_options_json=menu_options_json
        )
        if resolved_menu_id is None:
            conn.rollback()
            cursor.close()
            conn.close()
            return render_notice_page(
                "Menu Tidak Valid",
                "Pilihan menu belum lengkap. Silakan pilih ulang menu yang ingin disimpan.",
                back_url=url_for("create_reservation", date=reservation_date),
                back_label="Kembali",
                status_code=400
            )

        if special == "with_special":
            special_request = dish_desc
        else:
            special_request = None

        if fish_type in ("", "none"):
            fish_type = None

        if size == "":
            size = None

        if weight in ("", "0", "0.0"):
            weight = None

        resolved_fish_stock_id, resolved_special_stock_id = resolve_selected_stock_refs(
            cursor,
            reservation_date,
            resolved_menu_id,
            fish_type=fish_type,
            fish_weight=weight,
            fish_stock_id=fish_stock_id,
            special_stock_id=special_stock_id
        )

        is_valid_stock, stock_message = validate_stock_request(
            cursor,
            reservation_date,
            resolved_menu_id,
            qty,
            fish_type=fish_type,
            fish_size=size,
            fish_weight=weight,
            fish_stock_id=resolved_fish_stock_id,
            special_stock_id=resolved_special_stock_id,
            menu_options_json=menu_options_json
        )
        if not is_valid_stock:
            conn.rollback()
            cursor.close()
            conn.close()
            return render_notice_page(
                "Stock Tidak Cukup",
                stock_message,
                back_url=url_for("create_reservation", date=reservation_date),
                back_label="Kembali",
                status_code=400
            )

        cursor.execute("""
            INSERT INTO reservation_items
            (reservation_id, menu_id, quantity,
            special_request, dish_description,
            fish_type, fish_size, fish_weight, fish_stock_ref_id, special_stock_ref_id, menu_options_json)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """,(
            reservation_id,
            resolved_menu_id,
            qty,
            special_request,
            dish_desc,
            fish_type,
            size,
            weight,
            resolved_fish_stock_id,
            resolved_special_stock_id,
            menu_options_json or None
        ))
        log_reservation_history(
            cursor,
            reservation_id,
            "create",
            "menu",
            build_menu_history_summary(
                "create",
                get_menu_label_for_history(cursor, resolved_menu_id, menu_options_json=menu_options_json),
                qty=qty,
                note=build_menu_display_note(menu_options_json, dish_desc)
            ),
            actor=getattr(current_user, "id", "system"),
            reservation_item_id=cursor.lastrowid
        )

        reduce_stock_after_order(
            cursor,
            fish_stock_id=resolved_fish_stock_id,
            special_stock_id=resolved_special_stock_id,
            qty=int(qty)
        )

    conn.commit()

    cursor.close()
    conn.close()

    return redirect("/reservations")


# ================= RESERVATION MENU =================
@app.route("/reservation_menu/<int:reservation_id>")
@login_required
def reservation_menu(reservation_id):
    search_query = (request.args.get("search") or "").strip()

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    description_expr = dish_description_sql("ri")

    cursor.execute("""
    SELECT
        id,
        customer_name,
        table_number,
        people_count,
        reservation_datetime,
        description
    FROM reservations
    WHERE id = %s
    """,(reservation_id,))
    reservation_info = cursor.fetchone()

    if not reservation_info:
        cursor.close()
        conn.close()
        return render_notice_page(
            "Reservasi Tidak Ditemukan",
            "Data reservasi yang Anda buka sudah tidak tersedia.",
            back_url=url_for("reservations"),
            back_label="Kembali ke List Reservasi"
        )

    reservation_back_date = ""
    reservation_datetime = reservation_info.get("reservation_datetime")
    if isinstance(reservation_datetime, datetime):
        reservation_back_date = reservation_datetime.strftime("%Y-%m-%d")
    elif reservation_datetime:
        reservation_back_date = str(reservation_datetime).split(" ")[0]

    query = """
    SELECT
        ri.id,
        ri.reservation_id,
        ri.quantity,
        ri.menu_options_json,
        ri.fish_type,
        ri.fish_size,
        ri.fish_weight,
        COALESCE(fs.weight_unit, ds.weight_unit, 'ons') AS fish_weight_unit,
    """ + description_expr + """ AS dish_description,
        m.name,
        m.serving_type,
        m.price
    FROM reservation_items ri
    JOIN menus m ON ri.menu_id = m.id
    LEFT JOIN fish_stock fs ON fs.id = ri.fish_stock_ref_id
    LEFT JOIN daily_item_stock ds ON ds.id = ri.special_stock_ref_id
    WHERE ri.reservation_id = %s
    """
    params = [reservation_id]

    if search_query:
        query += """
        AND (
            m.name LIKE %s
            OR m.serving_type LIKE %s
            OR COALESCE(NULLIF(TRIM(ri.fish_type),''),'') LIKE %s
            OR COALESCE(NULLIF(TRIM(ri.fish_size),''),'') LIKE %s
            OR COALESCE(""" + description_expr + """,'') LIKE %s
        )
        """
        search_like = f"%{search_query}%"
        params.extend([search_like] * 5)

    query += " ORDER BY ri.id DESC"
    cursor.execute(query, params)

    reservation_menus = cursor.fetchall()
    for item in reservation_menus:
        item["effective_selected_options"] = get_effective_selected_options(item)
        item["fish_info"] = format_fish_info(item)
        item["price_display"] = format_currency(item.get("price"))
        item["display_menu_name"] = format_menu_label(
            get_payload_display_name(
                item.get("menu_options_json"),
                item.get("name"),
                item.get("serving_type")
            ),
            item.get("serving_type")
        )
        item["option_summary"] = build_option_summary_from_row(item)
        item["display_note"] = build_menu_display_note_from_row(item)

    cursor.close()
    conn.close()

    return render_template(
        "reservation_menu.html",
        reservation_menus=reservation_menus,
        reservation_id=reservation_id,
        reservation_info=reservation_info,
        reservation_back_date=reservation_back_date,
        search_query=search_query
    )
# ================= DASHBOARD =================

@app.route("/dashboard")
@login_required
def dashboard():

    selected_date = request.args.get("date")
    selected_divisi = request.args.get("divisi")
    search_query = (request.args.get("search") or "").strip()
    use_table_filters = request.args.get("use_table_filters") == "1"
    selected_tables = parse_table_filters(
        request.args.getlist("tables"),
        use_table_filters=use_table_filters
    )

    if not selected_date:
        selected_date = datetime.now().strftime("%Y-%m-%d")

    selected_columns = parse_dashboard_column_filters(request.args, selected_tables)
    selected_column_defs = get_selected_dashboard_column_defs(selected_tables, selected_columns)
    ordered_selected_tables = [
        table_key
        for table_key in DASHBOARD_TABLE_CONFIGS
        if table_key in selected_tables
    ]

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    data = get_daily_menu_recap_rows(cursor, selected_date, selected_divisi, search_query=search_query)

    fish_totals = get_daily_fish_totals(cursor, selected_date)
    menu_serving_totals = get_daily_menu_serving_totals(cursor, selected_date, selected_divisi)
    rice_requirement_rows = build_rice_requirement_rows(menu_serving_totals)
    fried_rice_rows = build_fried_rice_rows(menu_serving_totals)
    combined_menu_rows = build_combined_menu_rows(data)
    detailed_menu_rows = build_detailed_menu_rows(data)

    for index, row in enumerate(fish_totals, start=1):
        row["no"] = index

    dashboard_table_rows = build_dashboard_table_rows(
        fish_totals,
        rice_requirement_rows,
        fried_rice_rows,
        combined_menu_rows,
        detailed_menu_rows
    )

    cursor.close()
    conn.close()

    return render_template(
        "dashboard.html",
        data=data,
        fish_totals=fish_totals,
        rice_requirement_rows=rice_requirement_rows,
        fried_rice_rows=fried_rice_rows,
        combined_menu_rows=combined_menu_rows,
        detailed_menu_rows=detailed_menu_rows,
        selected_date=selected_date,
        selected_divisi=selected_divisi,
        search_query=search_query,
        selected_tables=selected_tables,
        dashboard_table_configs=DASHBOARD_TABLE_CONFIGS,
        dashboard_table_rows=dashboard_table_rows,
        selected_table_columns=selected_columns,
        selected_column_defs=selected_column_defs,
        ordered_selected_tables=ordered_selected_tables
    )
    
# ================= RESERVATION LIST =================

@app.route("/reservations")
@login_required
def reservations():

    selected_date = request.args.get("date")
    search_query = (request.args.get("search") or "").strip()

    if not selected_date:
        selected_date = datetime.now().strftime("%Y-%m-%d")

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    query = """
        SELECT id,customer_name,table_number,
               people_count,reservation_datetime,description
        FROM reservations
        WHERE DATE(reservation_datetime) = %s
    """
    params = [selected_date]

    if search_query:
        query += """
        AND (
            CAST(id AS CHAR) LIKE %s
            OR customer_name LIKE %s
            OR table_number LIKE %s
            OR COALESCE(description,'') LIKE %s
        )
        """
        search_like = f"%{search_query}%"
        params.extend([search_like] * 4)

    query += " ORDER BY id DESC"
    cursor.execute(query, params)

    data = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template(
        "reservations.html",
        data=data,
        selected_date=selected_date,
        search_query=search_query
    )


@app.route("/kitchen_live")
@login_required
def kitchen_live():
    selected_date = request.args.get("date")
    search_query = (request.args.get("search") or "").strip()

    if not selected_date:
        selected_date = datetime.now().strftime("%Y-%m-%d")

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    kitchen_reservations = get_kitchen_live_reservations(cursor, selected_date, search_query)
    cursor.close()
    conn.close()

    return render_template(
        "kitchen_live.html",
        kitchen_reservations=kitchen_reservations,
        selected_date=selected_date,
        search_query=search_query
    )


@app.route("/kitchen_live/save/<int:reservation_id>", methods=["POST"])
@login_required
def save_kitchen_live(reservation_id):
    selected_date = request.form.get("date") or datetime.now().strftime("%Y-%m-%d")
    search_query = (request.form.get("search") or "").strip()
    completed_item_ids = {
        int(item_id)
        for item_id in request.form.getlist("completed_item_ids")
        if str(item_id).strip().isdigit()
    }

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    ensure_kitchen_live_tables(cursor)

    cursor.execute("""
    SELECT customer_name
    FROM reservations
    WHERE id = %s
    """,(reservation_id,))
    reservation_row = cursor.fetchone()

    cursor.execute("""
    SELECT id
    FROM reservation_items
    WHERE reservation_id = %s
    ORDER BY id
    """,(reservation_id,))
    item_rows = cursor.fetchall()
    item_ids = [row["id"] for row in item_rows]

    if not reservation_row:
        cursor.close()
        conn.close()
        flash("Reservasi tidak ditemukan.")
        return redirect(url_for("kitchen_live", date=selected_date, search=search_query))

    if not item_ids:
        cursor.close()
        conn.close()
        flash(f"Reservasi {reservation_row['customer_name']} belum memiliki menu.")
        return redirect(url_for("kitchen_live", date=selected_date, search=search_query))

    for item_id in item_ids:
        is_completed = 1 if item_id in completed_item_ids else 0
        completed_at = datetime.now() if is_completed else None
        cursor.execute("""
        INSERT INTO reservation_item_completion (reservation_item_id, is_completed, completed_at)
        VALUES (%s, %s, %s)
        ON DUPLICATE KEY UPDATE
            is_completed = VALUES(is_completed),
            completed_at = VALUES(completed_at)
        """,(item_id, is_completed, completed_at))

    conn.commit()
    cursor.close()
    conn.close()

    if len(completed_item_ids) == len(item_ids):
        flash(f"Menu reservasi {reservation_row['customer_name']} sudah selesai.")
    else:
        flash(f"Progress kitchen untuk {reservation_row['customer_name']} berhasil disimpan.")

    return redirect(url_for("kitchen_live", date=selected_date, search=search_query))


@app.route("/kitchen_live/export")
@login_required
def export_kitchen_live_excel():
    selected_date = request.args.get("date")
    search_query = (request.args.get("search") or "").strip()

    if not selected_date:
        selected_date = datetime.now().strftime("%Y-%m-%d")

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    kitchen_reservations = get_kitchen_live_reservations(cursor, selected_date, search_query)
    cursor.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Kitchen Checklist"

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="2F75B5")
    thin_border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3")
    )
    white_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(color="FFFFFF", bold=True, size=16)

    ws.merge_cells("A1:J1")
    ws["A1"] = f"KITCHEN CHECKLIST - {selected_date}"
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["ID", "NAMA RESERVASI", "MEJA", "PAX", "WAKTU", "MENU", "DIVISI", "QTY", "KETERANGAN", "STATUS"]
    ws.append([])
    ws.append(headers)
    header_row = ws.max_row

    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for reservation in kitchen_reservations:
        if not reservation["menus"]:
            ws.append([
                reservation["reservation_id"],
                reservation["customer_name"],
                reservation["table_number"],
                reservation["people_count"],
                reservation["reservation_datetime"],
                "-",
                "-",
                0,
                reservation.get("reservation_description") or "-",
                "BELUM ADA MENU"
            ])
            continue

        for item in reservation["menus"]:
            note_parts = []
            if item.get("option_summary"):
                note_parts.append(item["option_summary"])
            if item.get("dish_description"):
                note_parts.append(item["dish_description"])
            if item.get("fish_info"):
                note_parts.append(item["fish_info"])
            ws.append([
                reservation["reservation_id"],
                reservation["customer_name"],
                reservation["table_number"],
                reservation["people_count"],
                reservation["reservation_datetime"],
                item.get("menu_display_name") or format_menu_label(item["menu_name"], item.get("serving_type")),
                item.get("divisi") or "-",
                item["quantity"],
                " | ".join(note_parts) or reservation.get("reservation_description") or "-",
                "SELESAI" if item["is_completed"] else "PROSES"
            ])

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, min_col=1, max_col=10):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    if ws.max_row > header_row:
        table = Table(
            displayName="KitchenLiveTable",
            ref=f"A{header_row}:J{ws.max_row}"
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showRowStripes=True,
            showColumnStripes=False
        )
        ws.add_table(table)

    for column_letter, width in {
        "A": 8,
        "B": 28,
        "C": 14,
        "D": 8,
        "E": 22,
        "F": 28,
        "G": 16,
        "H": 8,
        "I": 38,
        "J": 14
    }.items():
        ws.column_dimensions[column_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    file_date = selected_date.replace("-", "")
    return send_file(
        output,
        download_name=f"kitchen_checklist_{file_date}.xlsx",
        as_attachment=True
    )

# ================= DELETE RESERVATION =================

@app.route("/delete/<int:res_id>")
@login_required
def delete_reservation(res_id):

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    ensure_additional_stock_tables(cursor)

    cursor.execute("""
    SELECT
        ri.id,
        ri.reservation_id,
        ri.quantity,
        ri.fish_type,
        ri.fish_size,
        ri.fish_weight,
        ri.fish_stock_ref_id,
        ri.special_stock_ref_id,
        m.name,
        DATE_FORMAT(r.reservation_datetime, '%Y-%m-%d') AS reservation_date
    FROM reservation_items ri
    JOIN menus m ON m.id = ri.menu_id
    JOIN reservations r ON r.id = ri.reservation_id
    WHERE ri.reservation_id = %s
    """,(res_id,))
    items = cursor.fetchall()

    for item in items:
        restore_stock_for_item(cursor, item)

    log_reservation_history(
        cursor,
        res_id,
        "delete",
        "reservation",
        f"Reservasi dihapus bersama {len(items)} item menu.",
        actor=getattr(current_user, "id", "system")
    )

    cursor.execute("DELETE FROM reservations WHERE id=%s",(res_id,))
    conn.commit()

    cursor.close()
    conn.close()

    return redirect("/reservations")

# ================= DELETE ALL RESERVATIONS =================

@app.route("/delete_all_reservations")
@login_required
def delete_all_reservations():

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    ensure_additional_stock_tables(cursor)

    cursor.execute("""
    SELECT
        ri.id,
        ri.reservation_id,
        ri.quantity,
        ri.fish_type,
        ri.fish_size,
        ri.fish_weight,
        ri.fish_stock_ref_id,
        ri.special_stock_ref_id,
        m.name,
        DATE_FORMAT(r.reservation_datetime, '%Y-%m-%d') AS reservation_date
    FROM reservation_items ri
    JOIN menus m ON m.id = ri.menu_id
    JOIN reservations r ON r.id = ri.reservation_id
    """)
    items = cursor.fetchall()

    for item in items:
        restore_stock_for_item(cursor, item)

    cursor.execute("DELETE FROM reservation_items")
    cursor.execute("DELETE FROM reservations")

    conn.commit()

    cursor.close()
    conn.close()

    return redirect("/reservations")


# ================= EDIT RESERVATION =================

@app.route("/edit/<int:res_id>", methods=["GET","POST"])
@login_required
def edit_reservation(res_id):

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    if request.method == "POST":

        name = request.form["customer_name"]
        table = request.form["table_number"]
        people = request.form["people_count"]
        time = request.form["reservation_datetime"]
        deskripsi = request.form["description"]

        cursor.execute("SELECT * FROM reservations WHERE id=%s",(res_id,))
        existing_reservation = cursor.fetchone()

        cursor.execute("""
        UPDATE reservations
        SET customer_name=%s,
            table_number=%s,
            people_count=%s,
            reservation_datetime=%s,
            description=%s
        WHERE id=%s
        """,(name,table,people,time,deskripsi,res_id))

        if existing_reservation:
            log_reservation_history(
                cursor,
                res_id,
                "update",
                "reservation",
                f"Reservasi diubah dari {existing_reservation.get('customer_name')} / meja {existing_reservation.get('table_number')} menjadi {name} / meja {table}.",
                actor=getattr(current_user, "id", "system")
            )

        conn.commit()

        cursor.close()
        conn.close()

        return redirect("/reservations")

    cursor.execute("SELECT * FROM reservations WHERE id=%s",(res_id,))
    reservation = cursor.fetchone()

    if reservation and reservation.get("reservation_datetime"):
        reservation["reservation_datetime"] = reservation["reservation_datetime"].strftime("%Y-%m-%dT%H:%M")

    cursor.close()
    conn.close()

    return render_template("edit_reservation.html",reservation=reservation)

#================== EDIT MENU =================
@app.route("/edit_menu/<int:item_id>", methods=["GET","POST"])
@login_required
def edit_menu(item_id):

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    ensure_additional_stock_tables(cursor)

    # ================= AMBIL MENU YANG DIEDIT =================
    cursor.execute("""
        SELECT
            ri.id,
            ri.reservation_id,
            ri.menu_id,
            ri.quantity,
            ri.special_request,
            ri.dish_description,
            ri.menu_options_json,
            ri.fish_type,
            ri.fish_size,
            ri.fish_weight,
            ri.fish_stock_ref_id,
            ri.special_stock_ref_id,
            m.name,
            m.serving_type,
            DATE_FORMAT(r.reservation_datetime, '%Y-%m-%d') AS reservation_date
        FROM reservation_items ri
        JOIN menus m ON ri.menu_id = m.id
        JOIN reservations r ON r.id = ri.reservation_id
        WHERE ri.id = %s
    """,(item_id,))

    menu = cursor.fetchone()

    if not menu:
        cursor.close()
        conn.close()
        return render_notice_page(
            "Menu Tidak Ditemukan",
            "Menu yang ingin Anda edit sudah tidak ada. Silakan kembali ke halaman sebelumnya.",
            back_url=request.referrer or url_for("reservations"),
            back_label="Kembali",
            status_code=404
        )


    # ================= UPDATE MENU =================
    if request.method == "POST":

        qty = request.form["qty"]
        menu_id = request.form["menu_id"]
        display_menu_id = request.form.get("display_menu_id")

        fish_type = request.form.get("fish_type")
        fish_size = request.form.get("fish_size")
        fish_weight = request.form.get("fish_weight")
        fish_stock_id = request.form.get("fish_stock_id")
        special_stock_id = request.form.get("special_stock_id")

        special_request = request.form.get("special_request")
        dish_description = request.form.get("dish_description")
        menu_options_json = request.form.get("menu_options_json")
        menu_catalog, _, _ = get_stock_context(cursor, menu["reservation_date"])
        menu_options_json = build_menu_options_payload_from_form(
            request.form,
            menu_catalog,
            display_menu_id,
            fallback_payload=menu_options_json
        )
        resolved_menu_id = resolve_menu_submission(
            menu_catalog,
            menu_id,
            display_menu_id=display_menu_id,
            menu_options_json=menu_options_json
        )
        if resolved_menu_id is None:
            cursor.close()
            conn.close()
            return render_notice_page(
                "Menu Tidak Valid",
                "Pilihan menu belum lengkap. Silakan pilih ulang menu yang ingin disimpan.",
                back_url=url_for("edit_menu", item_id=item_id),
                back_label="Kembali",
                status_code=400
            )

        if special_request == "no_special":
            dish_description = None

        if fish_type in ("", "none"):
            fish_type = None

        if fish_size == "":
            fish_size = None

        if fish_weight in ("", "0", "0.0"):
            fish_weight = None

        resolved_fish_stock_id, resolved_special_stock_id = resolve_selected_stock_refs(
            cursor,
            menu["reservation_date"],
            resolved_menu_id,
            fish_type=fish_type,
            fish_weight=fish_weight,
            fish_stock_id=fish_stock_id,
            special_stock_id=special_stock_id
        )

        is_valid_stock, stock_message = validate_stock_request(
            cursor,
            menu["reservation_date"],
            resolved_menu_id,
            qty,
            fish_type=fish_type,
            fish_size=fish_size,
            fish_weight=fish_weight,
            fish_stock_id=resolved_fish_stock_id,
            special_stock_id=resolved_special_stock_id,
            current_item=menu,
            menu_options_json=menu_options_json
        )
        if not is_valid_stock:
            cursor.close()
            conn.close()
            return render_notice_page(
                "Stock Tidak Cukup",
                stock_message,
                back_url=url_for("edit_menu", item_id=item_id),
                back_label="Kembali",
                status_code=400
            )

        restore_stock_for_item(cursor, menu)

        cursor.execute("""
            UPDATE reservation_items
            SET quantity=%s,
                menu_id=%s,
                fish_type=%s,
            fish_size=%s,
            fish_weight=%s,
            fish_stock_ref_id=%s,
            special_stock_ref_id=%s,
            special_request=%s,
                dish_description=%s,
                menu_options_json=%s
            WHERE id=%s
        """,(
            qty,
            resolved_menu_id,
            fish_type,
            fish_size,
            fish_weight,
            resolved_fish_stock_id,
            resolved_special_stock_id,
            special_request,
            dish_description,
            menu_options_json or None,
            item_id
        ))
        previous_menu_label = get_menu_label_for_history(
            cursor,
            menu.get("menu_id"),
            menu_options_json=menu.get("menu_options_json"),
            fallback_name=menu.get("name"),
            fallback_serving_type=menu.get("serving_type")
        )
        updated_menu_label = get_menu_label_for_history(
            cursor,
            resolved_menu_id,
            menu_options_json=menu_options_json
        )
        log_reservation_history(
            cursor,
            menu["reservation_id"],
            "update",
            "menu",
            build_menu_history_summary(
                "update",
                updated_menu_label,
                qty=qty,
                note=build_menu_display_note(menu_options_json, dish_description),
                previous_menu_label=previous_menu_label,
                previous_qty=menu.get("quantity")
            ),
            actor=getattr(current_user, "id", "system"),
            reservation_item_id=item_id
        )

        reduce_stock_after_order(
            cursor,
            fish_stock_id=resolved_fish_stock_id,
            special_stock_id=resolved_special_stock_id,
            qty=int(qty)
        )

        conn.commit()
        cursor.close()
        conn.close()
        return redirect(url_for("edit_menu", item_id=item_id))


    # ================= CURRENT ORDER =================
    cursor.execute("""
        SELECT
            ri.id,
            ri.quantity,
            ri.special_request,
            ri.dish_description,
            ri.menu_options_json,
            ri.fish_type,
            ri.fish_size,
            ri.fish_weight,
            m.name,
            m.serving_type
        FROM reservation_items ri
        JOIN menus m ON ri.menu_id = m.id
        WHERE ri.reservation_id = %s
        ORDER BY ri.id DESC
        """,(menu["reservation_id"],))

    reservation_menus = cursor.fetchall()
    for item in reservation_menus:
        item["effective_selected_options"] = get_effective_selected_options(item)
        item["display_menu_name"] = format_menu_label(
            get_payload_display_name(
                item.get("menu_options_json"),
                item.get("name"),
                item.get("serving_type")
            ),
            item.get("serving_type")
        )
        item["option_summary"] = build_option_summary_from_row(item)
        item["display_note"] = build_menu_display_note_from_row(item)


    # ================= ALL MENU =================
    all_menus, nila_sizes, sea_fish = get_stock_context(cursor, menu["reservation_date"])
    tuna_stock_menus, tuna_piece_stock, rahang_tuna_stock = get_special_tuna_stock_context(cursor, menu["reservation_date"])

    nila_size_options = [row["size_category"] for row in nila_sizes]

    for fish in sea_fish:
        fish["display_weight"] = format_weight_display(fish["weight_ons"], fish.get("weight_unit"))
        fish["label"] = f"{fish['name'].capitalize()} - {fish['display_weight']} ({fish['fish_count']} stok)"

    menu["special_request_mode"] = "with_special" if menu["dish_description"] else "no_special"

    cursor.close()
    conn.close()

    return render_template(
        "edit_menu.html",
        menu=menu,
        reservation_menus=reservation_menus,
        all_menus=all_menus,
        reservation_id=menu["reservation_id"],
        nila_sizes=nila_size_options,
        sea_fish=sea_fish,
        tuna_piece_stock=tuna_piece_stock,
        rahang_tuna_stock=rahang_tuna_stock,
        selected_date=menu["reservation_date"]
    )


@app.route("/add_dish/<int:reservation_id>")
@login_required
def add_dish_page(reservation_id):

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    cursor.execute("""
        SELECT
            id,
            customer_name,
            table_number,
            DATE_FORMAT(reservation_datetime, '%Y-%m-%d') AS reservation_date
        FROM reservations
        WHERE id = %s
    """,(reservation_id,))
    reservation = cursor.fetchone()

    if not reservation:
        cursor.close()
        conn.close()
        return render_notice_page(
            "Reservasi Tidak Ditemukan",
            "Data reservasi ini sudah tidak tersedia. Silakan kembali dan pilih reservasi lain.",
            back_url=request.referrer or url_for("reservations"),
            back_label="Kembali",
            status_code=404
        )

    cursor.execute("""
        SELECT
            ri.id,
            ri.quantity,
            ri.special_request,
            ri.dish_description,
            ri.menu_options_json,
            ri.fish_type,
            ri.fish_size,
            ri.fish_weight,
            m.name,
            m.serving_type
        FROM reservation_items ri
        JOIN menus m ON ri.menu_id = m.id
        WHERE ri.reservation_id = %s
        ORDER BY ri.id DESC
    """,(reservation_id,))
    reservation_menus = cursor.fetchall()
    for item in reservation_menus:
        item["effective_selected_options"] = get_effective_selected_options(item)
        item["display_menu_name"] = format_menu_label(
            get_payload_display_name(
                item.get("menu_options_json"),
                item.get("name"),
                item.get("serving_type")
            ),
            item.get("serving_type")
        )
        item["option_summary"] = build_option_summary_from_row(item)
        item["display_note"] = build_menu_display_note_from_row(item)

    all_menus, nila_sizes, sea_fish = get_stock_context(cursor, reservation["reservation_date"])
    _, tuna_piece_stock, rahang_tuna_stock = get_special_tuna_stock_context(cursor, reservation["reservation_date"])

    nila_size_options = [row["size_category"] for row in nila_sizes]
    for fish in sea_fish:
        fish["display_weight"] = format_weight_display(fish["weight_ons"], fish.get("weight_unit"))
        fish["label"] = f"{fish['name'].capitalize()} - {fish['display_weight']} ({fish['fish_count']} stok)"

    cursor.close()
    conn.close()

    return render_template(
        "add_dish.html",
        reservation=reservation,
        reservation_id=reservation_id,
        reservation_menus=reservation_menus,
        all_menus=all_menus,
        nila_sizes=nila_size_options,
        sea_fish=sea_fish,
        tuna_piece_stock=tuna_piece_stock,
        rahang_tuna_stock=rahang_tuna_stock,
        selected_date=reservation["reservation_date"]
    )


# ================= DELETE MENU =================
@app.route("/delete_menu/<int:item_id>")
@login_required
def delete_menu(item_id):

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    ensure_additional_stock_tables(cursor)

    cursor.execute("""
    SELECT
        ri.id,
        ri.reservation_id,
        ri.quantity,
        ri.fish_type,
        ri.fish_size,
        ri.fish_weight,
        ri.fish_stock_ref_id,
        ri.special_stock_ref_id,
        m.name,
        DATE_FORMAT(r.reservation_datetime, '%Y-%m-%d') AS reservation_date
    FROM reservation_items ri
    JOIN menus m ON m.id = ri.menu_id
    JOIN reservations r ON r.id = ri.reservation_id
    WHERE ri.id = %s
    """,(item_id,))
    item = cursor.fetchone()

    if item:
        restore_stock_for_item(cursor, item)
        log_reservation_history(
            cursor,
            item["reservation_id"],
            "delete",
            "menu",
            f"Menu {item.get('name') or '-'} qty {item.get('quantity') or 0} dihapus.",
            actor=getattr(current_user, "id", "system"),
            reservation_item_id=item_id
        )
        cursor.execute("DELETE FROM reservation_items WHERE id=%s",(item_id,))

    conn.commit()

    cursor.close()
    conn.close()

    fallback_url = url_for("reservation_menu", reservation_id=item["reservation_id"]) if item else url_for("reservations")
    return redirect(request.referrer or fallback_url)

# ================= ADD MENU =================
@app.route("/add_menu/<int:reservation_id>", methods=["POST"])
@login_required
def add_menu(reservation_id):

    conn = get_db_connection()
    cursor = conn.cursor()
    ensure_additional_stock_tables(cursor)
    reservation_cursor = conn.cursor(dictionary=True)

    menu_id = request.form["menu_id"]
    display_menu_id = request.form.get("display_menu_id")
    qty = request.form["qty"]

    fish_type = request.form.get("fish_type")
    fish_size = request.form.get("fish_size")
    fish_weight = request.form.get("fish_weight")
    fish_stock_id = request.form.get("fish_stock_id")
    special_stock_id = request.form.get("special_stock_id")

    special_request = request.form.get("special_request", "no_special")
    dish_description = request.form.get("dish_description")
    menu_options_json = request.form.get("menu_options_json")

    reservation_cursor.execute("""
    SELECT DATE_FORMAT(reservation_datetime, '%Y-%m-%d') AS reservation_date
    FROM reservations
    WHERE id = %s
    """,(reservation_id,))
    reservation = reservation_cursor.fetchone()
    reservation_date = reservation["reservation_date"] if reservation else None
    menu_catalog, _, _ = get_stock_context(reservation_cursor, reservation_date)
    resolved_menu_id = resolve_menu_submission(
        menu_catalog,
        menu_id,
        display_menu_id=display_menu_id,
        menu_options_json=menu_options_json
    )
    if resolved_menu_id is None:
        conn.rollback()
        reservation_cursor.close()
        cursor.close()
        conn.close()
        return render_notice_page(
            "Menu Tidak Valid",
            "Pilihan menu belum lengkap. Silakan pilih ulang menu yang ingin disimpan.",
            back_url=url_for("add_dish_page", reservation_id=reservation_id),
            back_label="Kembali",
            status_code=400
        )

    if special_request == "no_special":
        dish_description = None

    if fish_type in ("", "none"):
        fish_type = None

    if fish_size == "":
        fish_size = None

    if fish_weight in ("", "0", "0.0"):
        fish_weight = None

    resolved_fish_stock_id, resolved_special_stock_id = resolve_selected_stock_refs(
        reservation_cursor,
        reservation_date,
        resolved_menu_id,
        fish_type=fish_type,
        fish_weight=fish_weight,
        fish_stock_id=fish_stock_id,
        special_stock_id=special_stock_id
    )

    is_valid_stock, stock_message = validate_stock_request(
        reservation_cursor,
        reservation_date,
        resolved_menu_id,
        qty,
        fish_type=fish_type,
        fish_size=fish_size,
        fish_weight=fish_weight,
        fish_stock_id=resolved_fish_stock_id,
        special_stock_id=resolved_special_stock_id,
        menu_options_json=menu_options_json
    )
    if not is_valid_stock:
        conn.rollback()
        reservation_cursor.close()
        cursor.close()
        conn.close()
        return render_notice_page(
            "Stock Tidak Cukup",
            stock_message,
            back_url=url_for("add_dish_page", reservation_id=reservation_id),
            back_label="Kembali",
            status_code=400
        )

    cursor.execute("""
        INSERT INTO reservation_items
        (reservation_id, menu_id, quantity, fish_type, fish_size, fish_weight, fish_stock_ref_id, special_stock_ref_id, special_request, dish_description, menu_options_json)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """,(
        reservation_id,
        resolved_menu_id,
        qty,
        fish_type,
        fish_size,
        fish_weight,
        resolved_fish_stock_id,
        resolved_special_stock_id,
        special_request,
        dish_description,
        menu_options_json or None
    ))
    log_reservation_history(
        cursor,
        reservation_id,
        "create",
        "menu",
        build_menu_history_summary(
            "create",
            get_menu_label_for_history(reservation_cursor, resolved_menu_id, menu_options_json=menu_options_json),
            qty=qty,
            note=build_menu_display_note(menu_options_json, dish_description)
        ),
        actor=getattr(current_user, "id", "system"),
        reservation_item_id=cursor.lastrowid
    )

    reduce_stock_after_order(
        cursor,
        fish_stock_id=resolved_fish_stock_id,
        special_stock_id=resolved_special_stock_id,
        qty=int(qty)
    )

    conn.commit()

    reservation_cursor.close()
    cursor.close()
    conn.close()

    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return "",204

    return redirect(url_for("add_dish_page", reservation_id=reservation_id))


# ================= CALCULATE =================
@app.route("/calculate")
@login_required
def calculate():
    use_table_filters = request.args.get("use_table_filters") == "1"
    selected_tables = parse_table_filters(
        request.args.getlist("tables"),
        use_table_filters=use_table_filters
    )

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)

    selected_date = request.args.get("date")

    if not selected_date:
        selected_date = date.today().strftime("%Y-%m-%d")

    # convert jika format DD/MM/YYYY
    try:
        selected_date = datetime.strptime(selected_date, "%d/%m/%Y").strftime("%Y-%m-%d")
    except:
        pass

    data = get_daily_menu_recap_rows(cursor, selected_date)
    fish_totals = get_daily_fish_totals(cursor, selected_date)
    menu_serving_totals = get_daily_menu_serving_totals(cursor, selected_date)
    rice_requirement_rows = build_rice_requirement_rows(menu_serving_totals)
    fried_rice_rows = build_fried_rice_rows(menu_serving_totals)
    combined_menu_rows = build_combined_menu_rows(data)
    detailed_menu_rows = build_detailed_menu_rows(data)
    total_today = sum(to_int_qty(row.get("total")) for row in data)

    for index, row in enumerate(fish_totals, start=1):
        row["no"] = index

    cursor.close()
    conn.close()

    return render_template(
        "calculate.html",
        data=data,
        total_today=total_today,
        fish_totals=fish_totals,
        rice_requirement_rows=rice_requirement_rows,
        fried_rice_rows=fried_rice_rows,
        combined_menu_rows=combined_menu_rows,
        detailed_menu_rows=detailed_menu_rows,
        selected_date=selected_date,
        selected_tables=selected_tables
    )
# ================= EXPORT EXCEL =================
@app.route("/export")
def export_excel():

    selected_date = request.args.get("date")
    selected_divisi = request.args.get("divisi")
    search_query = (request.args.get("search") or "").strip()
    use_table_filters = request.args.get("use_table_filters") == "1"
    selected_tables = parse_table_filters(
        request.args.getlist("tables"),
        use_table_filters=use_table_filters
    )
    selected_columns = parse_dashboard_column_filters(request.args, selected_tables)
    selected_column_defs = get_selected_dashboard_column_defs(selected_tables, selected_columns)
    ordered_selected_tables = [
        table_key
        for table_key in DASHBOARD_TABLE_CONFIGS
        if table_key in selected_tables
    ]

    if not selected_date:
        selected_date = datetime.now().strftime("%Y-%m-%d")

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    data = get_daily_menu_recap_rows(cursor, selected_date, selected_divisi, search_query=search_query)

    fish_totals = get_daily_fish_totals(cursor, selected_date)
    menu_serving_totals = get_daily_menu_serving_totals(cursor, selected_date, selected_divisi)

    rice_requirement_rows = build_rice_requirement_rows(menu_serving_totals)
    fried_rice_rows = build_fried_rice_rows(menu_serving_totals)
    combined_menu_rows = build_combined_menu_rows(data)
    detailed_menu_rows = build_detailed_menu_rows(data)
    for index, row in enumerate(fish_totals, start=1):
        row["no"] = index
    dashboard_table_rows = build_dashboard_table_rows(
        fish_totals,
        rice_requirement_rows,
        fried_rice_rows,
        combined_menu_rows,
        detailed_menu_rows
    )

    cursor.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Menu"
    title_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    header_fill = PatternFill("solid", fgColor="2F75B5")
    white_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(color="FFFFFF", bold=True, size=16)
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3")
    )

    def add_section_table(section_title, headers, rows, table_name, table_style_name, compact_number_cols=None):
        title_row = ws.max_row + 2
        end_col = len(headers)
        end_col_letter = get_column_letter(end_col)
        compact_number_cols = set(compact_number_cols or [])

        ws.merge_cells(
            start_row=title_row,
            start_column=1,
            end_row=title_row,
            end_column=end_col
        )
        title_cell = ws.cell(row=title_row, column=1, value=section_title)
        title_cell.fill = section_fill
        title_cell.font = Font(bold=True, size=13)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        header_row = title_row + 1
        for col_index, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_index, value=header)
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        safe_rows = rows or [[1] + ["-"] * (len(headers) - 1)]
        data_start_row = header_row + 1

        for row_index, row_values in enumerate(safe_rows, start=data_start_row):
            for col_index, cell_value in enumerate(row_values, start=1):
                cell = ws.cell(row=row_index, column=col_index, value=cell_value)
                cell.border = thin_border
                if col_index in compact_number_cols:
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                else:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

                if isinstance(cell_value, str):
                    normalized_value = normalize_text(cell_value)
                    if normalized_value.startswith("total") or normalized_value == "jumlah sambal nasgor":
                        cell.font = bold_font

        for col_index in compact_number_cols:
            if col_index <= end_col:
                column_letter = get_column_letter(col_index)
                ws.column_dimensions[column_letter].width = min(ws.column_dimensions[column_letter].width or 99, 10)

        data_end_row = data_start_row + len(safe_rows) - 1
        table = Table(
            displayName=table_name,
            ref=f"A{header_row}:{end_col_letter}{data_end_row}"
        )
        table.tableStyleInfo = TableStyleInfo(
            name=table_style_name,
            showRowStripes=True,
            showColumnStripes=False
        )
        ws.add_table(table)

    ws.merge_cells("A1:R1")
    ws["A1"] = f"REKAPAN MENU RAMADAN - {selected_date}"
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    for table_key in ordered_selected_tables:
        table_config = DASHBOARD_TABLE_CONFIGS.get(table_key, {})
        column_defs = selected_column_defs.get(table_key) or table_config.get("columns", [])
        headers = [column["label"] for column in column_defs]
        table_rows = dashboard_table_rows.get(table_key, [])
        export_rows = [
            [row.get(column["key"], "-") for column in column_defs]
            for row in table_rows
        ]
        compact_number_cols = {
            index
            for index, column in enumerate(column_defs, start=1)
            if column.get("key") in table_config.get("compact_number_keys", set())
        }

        add_section_table(
            table_config.get("excel_title") or table_config.get("title") or table_key,
            headers,
            export_rows,
            table_config.get("excel_table_name") or f"Table{len(headers)}",
            table_config.get("excel_style_name") or "TableStyleMedium2",
            compact_number_cols=compact_number_cols
        )

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 22
    ws.column_dimensions["I"].width = 24
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 18
    ws.column_dimensions["L"].width = 22
    ws.column_dimensions["M"].width = 22
    ws.column_dimensions["N"].width = 22
    ws.column_dimensions["O"].width = 16
    ws.column_dimensions["P"].width = 18
    ws.column_dimensions["Q"].width = 28
    ws.column_dimensions["R"].width = 10

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    file_date = selected_date.replace("-", "")

    return send_file(
        output,
        download_name=f"rekapan_menu_ramadan_{file_date}.xlsx",
        as_attachment=True
    )
#================== UPDATE STOCK =================
@app.route("/update_stock", methods=["GET","POST"])
@login_required
def update_stock():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    ensure_additional_stock_tables(cursor)

    selected_date = request.args.get("date")

    if not selected_date:
        selected_date = date.today().strftime("%Y-%m-%d")
    latest_menu_status_map = get_latest_menu_status_map(cursor, selected_date)
    latest_nila_status_map = get_latest_nila_status_map(cursor, selected_date)
    option_stock_map = get_latest_option_stock_map(cursor, selected_date)

    cursor.execute("""
        SELECT
            m.id,
            m.name,
            m.category,
            CASE
                WHEN LOWER(m.name) = 'rahang tuna' THEN 'tuna_weight'
                WHEN LOWER(m.name) IN ('dada tuna goreng','dada tuna bakar','paket dada tuna goreng','paket dada tuna bakar') THEN 'menu_piece'
                ELSE m.stock_type
            END AS stock_source,
            'ready' AS status
        FROM menus m
        ORDER BY m.category, m.name
    """)
    menus = cursor.fetchall()
    for menu in menus:
        latest_status_row = latest_menu_status_map.get(menu["id"])
        if latest_status_row:
            menu["status"] = row_value(latest_status_row, "status", 1, "ready") or "ready"

    tuna_stock_menus, tuna_piece_stock, rahang_tuna_stock = get_special_tuna_stock_context(cursor, selected_date)
    apply_special_stock_statuses(menus, tuna_piece_stock, rahang_tuna_stock)
    actor_name = getattr(current_user, "id", "system")

    if request.method == "POST":
        for m in menus:
            menu_name = (m.get("name") or "").strip().lower()
            stock_source = m.get("stock_source")
            if stock_source in ("menu_piece", "tuna_weight") or menu_name in SPECIAL_TUNA_MENU_NAMES:
                continue

            new_status = request.form.get(f"status_{m['id']}")
            if new_status is None:
                continue

            previous_status = m.get("status", "ready")

            cursor.execute("""
                INSERT INTO daily_menu_stock (menu_id, status, stock_date)
                VALUES (%s,%s,%s)
                ON DUPLICATE KEY UPDATE status=%s
            """,(m["id"], new_status, selected_date, new_status))
            log_stock_history(cursor, "menu_status", m["name"], previous_status, new_status, actor=actor_name)

        for stock_menu in tuna_stock_menus:
            qty_value = request.form.get(f"piece_qty_{stock_menu['id']}", "").strip()
            status_value = request.form.get(f"piece_status_{stock_menu['id']}", "ready")
            qty_number = int(qty_value) if qty_value else 0
            status_value = effective_stock_status(status_value, qty_number)
            current_piece_stock = tuna_piece_stock.get(stock_menu["id"], {})

            cursor.execute("""
                INSERT INTO daily_item_stock (menu_id, stock_date, weight_ons, available_qty, status)
                VALUES (%s,%s,0,%s,%s)
                ON DUPLICATE KEY UPDATE available_qty=%s, status=%s
            """,(stock_menu["id"], selected_date, qty_number, status_value, qty_number, status_value))
            log_stock_history(
                cursor,
                "special_stock_qty",
                "Paket Dada Tuna",
                current_piece_stock.get("available_qty"),
                qty_number,
                actor=actor_name
            )
            log_stock_history(
                cursor,
                "special_stock_status",
                "Paket Dada Tuna",
                current_piece_stock.get("status"),
                status_value,
                actor=actor_name
            )

            for package_menu_name in PACKAGE_TUNA_MENU_NAMES:
                cursor.execute("""
                    INSERT INTO daily_menu_stock (menu_id, status, stock_date)
                    SELECT id, %s, %s
                    FROM menus
                    WHERE LOWER(name) = %s
                    ON DUPLICATE KEY UPDATE status = VALUES(status)
                """,(status_value, selected_date, package_menu_name))

        nila_sizes = ["kecil","sedang","besar","jumbo","super_jumbo"]
        for size in nila_sizes:
            status = request.form.get(f"nila_status_{size}")
            if not status:
                continue
            existing = latest_nila_status_map.get(size)
            cursor.execute("""
                SELECT id
                FROM fish_stock
                WHERE fish_type_id = 4
                AND size_category = %s
                AND stock_date = %s
                ORDER BY id DESC
                LIMIT 1
            """,(size, selected_date))
            current_date_row = cursor.fetchone()
            if current_date_row:
                cursor.execute("""
                    UPDATE fish_stock
                    SET status = %s
                    WHERE id = %s
                """,(status, current_date_row["id"]))
            else:
                cursor.execute("""
                    INSERT INTO fish_stock
                    (fish_type_id,size_category,status,stock_date)
                    VALUES (4,%s,%s,%s)
                """,(size,status,selected_date))
            log_stock_history(
                cursor,
                "nila_status",
                f"Nila {size}",
                row_value(existing, "status", 2, "ready"),
                status,
                actor=actor_name
            )

        for option_key, option_values in OPTION_STOCK_LABELS.items():
            for option_value, option_label in option_values.items():
                option_status = request.form.get(f"option_stock_{option_key}_{option_value}")
                if not option_status:
                    continue

                previous_status = option_stock_map.get((option_key, option_value), "ready")
                cursor.execute("""
                    INSERT INTO menu_option_stock (option_key, option_value, status, stock_date)
                    VALUES (%s,%s,%s,%s)
                    ON DUPLICATE KEY UPDATE status = VALUES(status)
                """,(option_key, option_value, option_status, selected_date))
                log_stock_history(
                    cursor,
                    "menu_option",
                    f"{option_key}:{option_label}",
                    previous_status,
                    option_status,
                    actor=actor_name
                )

        conn.commit()
        cursor.close()
        conn.close()
        return redirect(f"/update_stock?date={selected_date}")

    fish_stock_today = [
        row
        for row in get_latest_sea_fish_stock_rows(cursor, selected_date)
        if float(row.get("weight_ons") or 0) > 0 or int(row.get("fish_count") or 0) > 0
    ]

    for fish in fish_stock_today:
        fish["display_weight"] = format_weight_display(fish["weight_ons"], fish.get("weight_unit"))

    nila_status = {
        size_key: row_value(latest_nila_status_map.get(size_key), "status", 2, "ready")
        for size_key in ["kecil","sedang","besar","jumbo","super_jumbo"]
    }
    nila_indicator = {}
    for s in ["kecil","sedang","besar","jumbo","super_jumbo"]:
        nila_status.setdefault(s,"ready")
        nila_indicator[s] = status_indicator(nila_status[s])

    tuna_piece_stock_by_menu = {}
    for stock_menu in tuna_stock_menus:
        tuna_piece_stock_by_menu[stock_menu["id"]] = tuna_piece_stock.get(stock_menu["id"], {
            "available_qty": 0,
            "status": "not_ready",
            "display_name": "Paket Dada Tuna",
            "status_dot": status_indicator("not_ready")
        })

    rahang_tuna_rows = []
    for row in rahang_tuna_stock.get("all_rows", []):
        if row.get("weight_ons") and float(row["weight_ons"]) > 0:
            row["display_weight"] = format_weight_display(row["weight_ons"], row.get("weight_unit"))
            rahang_tuna_rows.append(row)

    option_stock_rows = []
    for option_key, option_values in OPTION_STOCK_LABELS.items():
        option_group_label = "Bumbu Taiwan Snack" if option_key == "seasoning" else "Topping Taiwan Snack"
        for option_value, option_label in option_values.items():
            current_status = option_stock_map.get((option_key, option_value), "ready")
            option_stock_rows.append({
                "group_label": option_group_label,
                "option_key": option_key,
                "option_value": option_value,
                "option_label": option_label,
                "status": current_status
            })

    cursor.close()
    conn.close()

    return render_template(
        "update_stock.html",
        menus=menus,
        fish_stock_today=fish_stock_today,
        nila_status=nila_status,
        nila_indicator=nila_indicator,
        tuna_stock_menus=tuna_stock_menus,
        tuna_piece_stock=tuna_piece_stock_by_menu,
        rahang_tuna_rows=rahang_tuna_rows,
        rahang_tuna_summary=rahang_tuna_stock.get("summary", {}),
        option_stock_rows=option_stock_rows,
        selected_date=selected_date
    )


#================== CLEAR FISH STOCK =================
@app.route("/clear_fish_stock")
@login_required
def clear_fish_stock():

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
    DELETE FROM fish_stock
    WHERE stock_date = CURDATE()
    """)

    conn.commit()

    cursor.close()
    conn.close()

    return redirect("/update_stock")


#================== FISH STOCK =================
@app.route("/fish_stock", methods=["GET","POST"])
@login_required
def fish_stock():
    selected_date = request.values.get("date")

    if not selected_date:
        selected_date = date.today().strftime("%Y-%m-%d")

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    ensure_additional_stock_tables(cursor)
    _, _, rahang_tuna_stock = get_special_tuna_stock_context(cursor, selected_date)
    rahang_menu = rahang_tuna_stock.get("menu")

    if request.method == "POST":
        stock_kind = request.form.get("stock_kind", "sea")

        if stock_kind == "sea":
            fish_type_id = request.form.get("fish_type_id")
            weight_unit = request.form.get("weight_unit", "ons")
            weight = normalize_weight_to_ons(request.form.get("weight_ons"), weight_unit)
            count = request.form.get("fish_count") or 0
            status = request.form.get("status", "ready")

            if fish_type_id:
                cursor.execute("""
                INSERT INTO fish_stock
                (fish_type_id, weight_ons, weight_unit, fish_count, status, stock_date)
                VALUES (%s,%s,%s,%s,%s,%s)
                """,(fish_type_id, weight, weight_unit, count, status, selected_date))

        elif stock_kind == "rahang_tuna" and rahang_menu:
            weight_unit = request.form.get("rahang_weight_unit", "ons")
            weight = normalize_weight_to_ons(request.form.get("rahang_weight"), weight_unit)
            qty = request.form.get("rahang_qty") or 0
            status = request.form.get("rahang_status", "ready")

            cursor.execute("""
                INSERT INTO daily_item_stock (menu_id, stock_date, weight_ons, weight_unit, available_qty, status)
                VALUES (%s,%s,%s,%s,%s,%s)
                ON DUPLICATE KEY UPDATE available_qty=%s, status=%s, weight_unit=%s
            """,(rahang_menu["id"], selected_date, weight, weight_unit, qty, status, qty, status, weight_unit))

        conn.commit()
        cursor.close()
        conn.close()
        return redirect(url_for("fish_stock", date=selected_date))

    cursor.execute("""
        SELECT
            id,
            name
        FROM fish_types ft
        WHERE ft.fish_category = 'sea'
        ORDER BY ft.name
    """)
    sea_fish_inputs = cursor.fetchall()

    fish_stock_today = get_latest_sea_fish_stock_rows(cursor, selected_date)

    for fish in fish_stock_today:
        fish["display_weight"] = format_weight_display(fish["weight_ons"], fish.get("weight_unit"))

    rahang_stock_rows = list(rahang_tuna_stock.get("all_rows", [])) if rahang_menu else []

    for item in rahang_stock_rows:
        item["display_weight"] = format_weight_display(item["weight_ons"], item.get("weight_unit"))

    cursor.close()
    conn.close()

    return render_template(
        "fish_stock.html",
        fish_stock_today=fish_stock_today,
        sea_fish_inputs=sea_fish_inputs,
        rahang_tuna_stock=rahang_stock_rows,
        rahang_tuna_summary=rahang_tuna_stock.get("summary", {}),
        rahang_menu=rahang_menu,
        selected_date=selected_date
    )


@app.route("/fish_stock/update/<int:stock_id>", methods=["POST"])
@login_required
def update_fish_stock_entry(stock_id):

    selected_date = request.form.get("date") or date.today().strftime("%Y-%m-%d")

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
    UPDATE fish_stock
    SET weight_ons=%s,
        weight_unit=%s,
        fish_count=%s,
        status=%s
    WHERE id=%s
    """,(
        normalize_weight_to_ons(request.form.get("weight_ons"), request.form.get("weight_unit", "ons")),
        request.form.get("weight_unit", "ons"),
        request.form.get("fish_count"),
        request.form.get("status", "ready"),
        stock_id
    ))

    conn.commit()
    cursor.close()
    conn.close()

    return redirect(url_for("fish_stock", date=selected_date))


@app.route("/fish_stock/delete/<int:stock_id>")
@login_required
def delete_fish_stock_entry(stock_id):

    selected_date = request.args.get("date") or date.today().strftime("%Y-%m-%d")

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("DELETE FROM fish_stock WHERE id=%s",(stock_id,))
    conn.commit()

    cursor.close()
    conn.close()

    return redirect(url_for("fish_stock", date=selected_date))


@app.route("/tuna_stock/update/<int:stock_id>", methods=["POST"])
@login_required
def update_tuna_stock_entry(stock_id):

    selected_date = request.form.get("date") or date.today().strftime("%Y-%m-%d")

    conn = get_db_connection()
    cursor = conn.cursor()
    ensure_additional_stock_tables(cursor)

    cursor.execute("""
    UPDATE daily_item_stock
    SET weight_ons=%s,
        weight_unit=%s,
        available_qty=%s,
        status=%s
    WHERE id=%s
    """,(
        normalize_weight_to_ons(request.form.get("weight_ons", 0), request.form.get("weight_unit", "ons")),
        request.form.get("weight_unit", "ons"),
        request.form.get("available_qty", 0),
        request.form.get("status", "ready"),
        stock_id
    ))

    conn.commit()
    cursor.close()
    conn.close()

    return redirect(url_for("fish_stock", date=selected_date))


@app.route("/tuna_stock/delete/<int:stock_id>")
@login_required
def delete_tuna_stock_entry(stock_id):

    selected_date = request.args.get("date") or date.today().strftime("%Y-%m-%d")

    conn = get_db_connection()
    cursor = conn.cursor()
    ensure_additional_stock_tables(cursor)

    cursor.execute("DELETE FROM daily_item_stock WHERE id=%s",(stock_id,))
    conn.commit()

    cursor.close()
    conn.close()

    return redirect(url_for("fish_stock", date=selected_date))


@app.route("/save_fish_stock", methods=["POST"])
@login_required
def save_fish_stock():

    data = request.get_json()

    conn = get_db_connection()
    cursor = conn.cursor()

    for f in data["fish"]:

        if not f["weight"] or not f["count"]:
            continue

        cursor.execute("""
        INSERT INTO fish_stock
        (fish_type_id, weight_ons, fish_count, status, stock_date)
        VALUES (
            (SELECT id FROM fish_types WHERE name=%s),
            %s,
            %s,
            'ready',
            CURDATE()
        )
        """,(f["name"], f["weight"], f["count"]))

    conn.commit()

    cursor.close()
    conn.close()

    return {"status":"ok"}


@app.route("/stock_history")
@login_required
def stock_history():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    ensure_additional_stock_tables(cursor)
    cursor.execute("""
        SELECT id, stock_scope, target_name, previous_value, new_value, actor_name, notes, created_at
        FROM stock_change_log
        ORDER BY created_at DESC, id DESC
        LIMIT 300
    """)
    history_rows = cursor.fetchall()
    cursor.close()
    conn.close()

    return render_template("stock_history.html", history_rows=history_rows)


@app.route("/reservation_history")
@login_required
def reservation_history():
    reservation_id = request.args.get("reservation_id")
    action_type = (request.args.get("action_type") or "").strip().lower()
    allowed_action_types = {"create", "update", "delete"}
    selected_action_type = action_type if action_type in allowed_action_types else None

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    ensure_additional_stock_tables(cursor)

    query = """
            SELECT id, reservation_id, reservation_item_id, action_type, change_scope, summary, actor_name, created_at
            FROM reservation_change_log
    """
    filters = []
    params = []

    if reservation_id and str(reservation_id).isdigit():
        filters.append("reservation_id = %s")
        params.append(reservation_id)

    if selected_action_type:
        filters.append("action_type = %s")
        params.append(selected_action_type)

    if filters:
        query += " WHERE " + " AND ".join(filters)

    query += """
        ORDER BY created_at DESC, id DESC
        LIMIT 300
    """

    cursor.execute(query, tuple(params))

    history_rows = cursor.fetchall()
    history_rows = prepare_reservation_history_rows(cursor, history_rows)
    cursor.close()
    conn.close()

    return render_template(
        "reservation_history.html",
        history_rows=history_rows,
        reservation_id=reservation_id,
        action_type_filter=selected_action_type
    )
if __name__ == "__main__":
    app.run(debug=True)

