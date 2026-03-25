import io
from datetime import date, datetime

from flask import flash, redirect, render_template, request, send_file, url_for
from flask_login import login_required
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from db import get_db_connection
from services.common import format_menu_label, normalize_text
from services.reporting import (
    DASHBOARD_TABLE_CONFIGS,
    build_combined_menu_rows,
    build_dashboard_table_rows,
    build_detailed_menu_rows,
    build_fried_rice_rows,
    build_rice_requirement_rows,
    ensure_kitchen_live_tables,
    get_daily_fish_totals,
    get_daily_menu_recap_rows,
    get_daily_menu_serving_totals,
    get_kitchen_live_reservations,
    get_selected_dashboard_column_defs,
    parse_dashboard_column_filters,
    parse_table_filters,
    to_int_qty,
)


def register_dashboard_routes(app):
    @app.route("/dashboard")
    @login_required
    def dashboard():

        selected_date = request.args.get("date")
        selected_divisi = request.args.get("divisi")
        search_query = (request.args.get("search") or "").strip()
        use_table_filters = request.args.get("use_table_filters") == "1"
        selected_tables = parse_table_filters(
            request.args.getlist("tables"),
            use_table_filters=use_table_filters
        )

        if not selected_date:
            selected_date = datetime.now().strftime("%Y-%m-%d")

        selected_columns = parse_dashboard_column_filters(request.args, selected_tables)
        selected_column_defs = get_selected_dashboard_column_defs(selected_tables, selected_columns)
        ordered_selected_tables = [
            table_key
            for table_key in DASHBOARD_TABLE_CONFIGS
            if table_key in selected_tables
        ]

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        data = get_daily_menu_recap_rows(cursor, selected_date, selected_divisi, search_query=search_query)

        fish_totals = get_daily_fish_totals(cursor, selected_date)
        menu_serving_totals = get_daily_menu_serving_totals(cursor, selected_date, selected_divisi)
        rice_requirement_rows = build_rice_requirement_rows(menu_serving_totals)
        fried_rice_rows = build_fried_rice_rows(menu_serving_totals)
        combined_menu_rows = build_combined_menu_rows(data)
        detailed_menu_rows = build_detailed_menu_rows(data)

        for index, row in enumerate(fish_totals, start=1):
            row["no"] = index

        dashboard_table_rows = build_dashboard_table_rows(
            fish_totals,
            rice_requirement_rows,
            fried_rice_rows,
            combined_menu_rows,
            detailed_menu_rows
        )

        cursor.close()
        conn.close()

        return render_template(
            "dashboard.html",
            data=data,
            fish_totals=fish_totals,
            rice_requirement_rows=rice_requirement_rows,
            fried_rice_rows=fried_rice_rows,
            combined_menu_rows=combined_menu_rows,
            detailed_menu_rows=detailed_menu_rows,
            selected_date=selected_date,
            selected_divisi=selected_divisi,
            search_query=search_query,
            selected_tables=selected_tables,
            dashboard_table_configs=DASHBOARD_TABLE_CONFIGS,
            dashboard_table_rows=dashboard_table_rows,
            selected_table_columns=selected_columns,
            selected_column_defs=selected_column_defs,
            ordered_selected_tables=ordered_selected_tables
        )
    
    @app.route("/kitchen_live")
    @login_required
    def kitchen_live():
        selected_date = request.args.get("date")
        search_query = (request.args.get("search") or "").strip()

        if not selected_date:
            selected_date = datetime.now().strftime("%Y-%m-%d")

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        kitchen_reservations = get_kitchen_live_reservations(cursor, selected_date, search_query)
        cursor.close()
        conn.close()

        return render_template(
            "kitchen_live.html",
            kitchen_reservations=kitchen_reservations,
            selected_date=selected_date,
            search_query=search_query
        )


    @app.route("/kitchen_live/save/<int:reservation_id>", methods=["POST"])
    @login_required
    def save_kitchen_live(reservation_id):
        selected_date = request.form.get("date") or datetime.now().strftime("%Y-%m-%d")
        search_query = (request.form.get("search") or "").strip()
        completed_item_ids = {
            int(item_id)
            for item_id in request.form.getlist("completed_item_ids")
            if str(item_id).strip().isdigit()
        }

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        ensure_kitchen_live_tables(cursor)

        cursor.execute("""
        SELECT customer_name
        FROM reservations
        WHERE id = %s
        """,(reservation_id,))
        reservation_row = cursor.fetchone()

        cursor.execute("""
        SELECT id
        FROM reservation_items
        WHERE reservation_id = %s
        ORDER BY id
        """,(reservation_id,))
        item_rows = cursor.fetchall()
        item_ids = [row["id"] for row in item_rows]

        if not reservation_row:
            cursor.close()
            conn.close()
            flash("Reservasi tidak ditemukan.")
            return redirect(url_for("kitchen_live", date=selected_date, search=search_query))

        if not item_ids:
            cursor.close()
            conn.close()
            flash(f"Reservasi {reservation_row['customer_name']} belum memiliki menu.")
            return redirect(url_for("kitchen_live", date=selected_date, search=search_query))

        for item_id in item_ids:
            is_completed = 1 if item_id in completed_item_ids else 0
            completed_at = datetime.now() if is_completed else None
            cursor.execute("""
            INSERT INTO reservation_item_completion (reservation_item_id, is_completed, completed_at)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE
                is_completed = VALUES(is_completed),
                completed_at = VALUES(completed_at)
            """,(item_id, is_completed, completed_at))

        conn.commit()
        cursor.close()
        conn.close()

        if len(completed_item_ids) == len(item_ids):
            flash(f"Menu reservasi {reservation_row['customer_name']} sudah selesai.")
        else:
            flash(f"Progress kitchen untuk {reservation_row['customer_name']} berhasil disimpan.")

        return redirect(url_for("kitchen_live", date=selected_date, search=search_query))


    @app.route("/kitchen_live/export")
    @login_required
    def export_kitchen_live_excel():
        selected_date = request.args.get("date")
        search_query = (request.args.get("search") or "").strip()

        if not selected_date:
            selected_date = datetime.now().strftime("%Y-%m-%d")

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        kitchen_reservations = get_kitchen_live_reservations(cursor, selected_date, search_query)
        cursor.close()
        conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Kitchen Checklist"

        title_fill = PatternFill("solid", fgColor="1F4E78")
        header_fill = PatternFill("solid", fgColor="2F75B5")
        thin_border = Border(
            left=Side(style="thin", color="D9E2F3"),
            right=Side(style="thin", color="D9E2F3"),
            top=Side(style="thin", color="D9E2F3"),
            bottom=Side(style="thin", color="D9E2F3")
        )
        white_font = Font(color="FFFFFF", bold=True, size=12)
        title_font = Font(color="FFFFFF", bold=True, size=16)

        ws.merge_cells("A1:J1")
        ws["A1"] = f"KITCHEN CHECKLIST - {selected_date}"
        ws["A1"].fill = title_fill
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        headers = ["ID", "NAMA RESERVASI", "MEJA", "PAX", "WAKTU", "MENU", "DIVISI", "QTY", "KETERANGAN", "STATUS"]
        ws.append([])
        ws.append(headers)
        header_row = ws.max_row

        for cell in ws[header_row]:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for reservation in kitchen_reservations:
            if not reservation["menus"]:
                ws.append([
                    reservation["reservation_id"],
                    reservation["customer_name"],
                    reservation["table_number"],
                    reservation["people_count"],
                    reservation["reservation_datetime"],
                    "-",
                    "-",
                    0,
                    reservation.get("reservation_description") or "-",
                    "BELUM ADA MENU"
                ])
                continue

            for item in reservation["menus"]:
                note_parts = []
                if item.get("option_summary"):
                    note_parts.append(item["option_summary"])
                if item.get("dish_description"):
                    note_parts.append(item["dish_description"])
                if item.get("fish_info"):
                    note_parts.append(item["fish_info"])
                ws.append([
                    reservation["reservation_id"],
                    reservation["customer_name"],
                    reservation["table_number"],
                    reservation["people_count"],
                    reservation["reservation_datetime"],
                    item.get("menu_display_name") or format_menu_label(item["menu_name"], item.get("serving_type")),
                    item.get("divisi") or "-",
                    item["quantity"],
                    " | ".join(note_parts) or reservation.get("reservation_description") or "-",
                    "SELESAI" if item["is_completed"] else "PROSES"
                ])

        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, min_col=1, max_col=10):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        if ws.max_row > header_row:
            table = Table(
                displayName="KitchenLiveTable",
                ref=f"A{header_row}:J{ws.max_row}"
            )
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showRowStripes=True,
                showColumnStripes=False
            )
            ws.add_table(table)

        for column_letter, width in {
            "A": 8,
            "B": 28,
            "C": 14,
            "D": 8,
            "E": 22,
            "F": 28,
            "G": 16,
            "H": 8,
            "I": 38,
            "J": 14
        }.items():
            ws.column_dimensions[column_letter].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        file_date = selected_date.replace("-", "")
        return send_file(
            output,
            download_name=f"kitchen_checklist_{file_date}.xlsx",
            as_attachment=True
        )


    @app.route("/calculate")
    @login_required
    def calculate():
        use_table_filters = request.args.get("use_table_filters") == "1"
        selected_tables = parse_table_filters(
            request.args.getlist("tables"),
            use_table_filters=use_table_filters
        )

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        selected_date = request.args.get("date")

        if not selected_date:
            selected_date = date.today().strftime("%Y-%m-%d")

        # convert jika format DD/MM/YYYY
        try:
            selected_date = datetime.strptime(selected_date, "%d/%m/%Y").strftime("%Y-%m-%d")
        except:
            pass

        data = get_daily_menu_recap_rows(cursor, selected_date)
        fish_totals = get_daily_fish_totals(cursor, selected_date)
        menu_serving_totals = get_daily_menu_serving_totals(cursor, selected_date)
        rice_requirement_rows = build_rice_requirement_rows(menu_serving_totals)
        fried_rice_rows = build_fried_rice_rows(menu_serving_totals)
        combined_menu_rows = build_combined_menu_rows(data)
        detailed_menu_rows = build_detailed_menu_rows(data)
        total_today = sum(to_int_qty(row.get("total")) for row in data)

        for index, row in enumerate(fish_totals, start=1):
            row["no"] = index

        cursor.close()
        conn.close()

        return render_template(
            "calculate.html",
            data=data,
            total_today=total_today,
            fish_totals=fish_totals,
            rice_requirement_rows=rice_requirement_rows,
            fried_rice_rows=fried_rice_rows,
            combined_menu_rows=combined_menu_rows,
            detailed_menu_rows=detailed_menu_rows,
            selected_date=selected_date,
            selected_tables=selected_tables
        )
    # ================= EXPORT EXCEL =================
    @app.route("/export")
    def export_excel():

        selected_date = request.args.get("date")
        selected_divisi = request.args.get("divisi")
        search_query = (request.args.get("search") or "").strip()
        use_table_filters = request.args.get("use_table_filters") == "1"
        selected_tables = parse_table_filters(
            request.args.getlist("tables"),
            use_table_filters=use_table_filters
        )
        selected_columns = parse_dashboard_column_filters(request.args, selected_tables)
        selected_column_defs = get_selected_dashboard_column_defs(selected_tables, selected_columns)
        ordered_selected_tables = [
            table_key
            for table_key in DASHBOARD_TABLE_CONFIGS
            if table_key in selected_tables
        ]

        if not selected_date:
            selected_date = datetime.now().strftime("%Y-%m-%d")

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        data = get_daily_menu_recap_rows(cursor, selected_date, selected_divisi, search_query=search_query)

        fish_totals = get_daily_fish_totals(cursor, selected_date)
        menu_serving_totals = get_daily_menu_serving_totals(cursor, selected_date, selected_divisi)

        rice_requirement_rows = build_rice_requirement_rows(menu_serving_totals)
        fried_rice_rows = build_fried_rice_rows(menu_serving_totals)
        combined_menu_rows = build_combined_menu_rows(data)
        detailed_menu_rows = build_detailed_menu_rows(data)
        for index, row in enumerate(fish_totals, start=1):
            row["no"] = index
        dashboard_table_rows = build_dashboard_table_rows(
            fish_totals,
            rice_requirement_rows,
            fried_rice_rows,
            combined_menu_rows,
            detailed_menu_rows
        )

        cursor.close()
        conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = "Rekap Menu"
        title_fill = PatternFill("solid", fgColor="1F4E78")
        section_fill = PatternFill("solid", fgColor="D9EAF7")
        header_fill = PatternFill("solid", fgColor="2F75B5")
        white_font = Font(color="FFFFFF", bold=True, size=12)
        title_font = Font(color="FFFFFF", bold=True, size=16)
        bold_font = Font(bold=True)
        thin_border = Border(
            left=Side(style="thin", color="D9E2F3"),
            right=Side(style="thin", color="D9E2F3"),
            top=Side(style="thin", color="D9E2F3"),
            bottom=Side(style="thin", color="D9E2F3")
        )

        def add_section_table(section_title, headers, rows, table_name, table_style_name, compact_number_cols=None):
            title_row = ws.max_row + 2
            end_col = len(headers)
            end_col_letter = get_column_letter(end_col)
            compact_number_cols = set(compact_number_cols or [])

            ws.merge_cells(
                start_row=title_row,
                start_column=1,
                end_row=title_row,
                end_column=end_col
            )
            title_cell = ws.cell(row=title_row, column=1, value=section_title)
            title_cell.fill = section_fill
            title_cell.font = Font(bold=True, size=13)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")

            header_row = title_row + 1
            for col_index, header in enumerate(headers, start=1):
                cell = ws.cell(row=header_row, column=col_index, value=header)
                cell.fill = header_fill
                cell.font = white_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            safe_rows = rows or [[1] + ["-"] * (len(headers) - 1)]
            data_start_row = header_row + 1

            for row_index, row_values in enumerate(safe_rows, start=data_start_row):
                for col_index, cell_value in enumerate(row_values, start=1):
                    cell = ws.cell(row=row_index, column=col_index, value=cell_value)
                    cell.border = thin_border
                    if col_index in compact_number_cols:
                        cell.alignment = Alignment(horizontal="center", vertical="top")
                    else:
                        cell.alignment = Alignment(vertical="top", wrap_text=True)

                    if isinstance(cell_value, str):
                        normalized_value = normalize_text(cell_value)
                        if normalized_value.startswith("total") or normalized_value == "jumlah sambal nasgor":
                            cell.font = bold_font

            for col_index in compact_number_cols:
                if col_index <= end_col:
                    column_letter = get_column_letter(col_index)
                    ws.column_dimensions[column_letter].width = min(ws.column_dimensions[column_letter].width or 99, 10)

            data_end_row = data_start_row + len(safe_rows) - 1
            table = Table(
                displayName=table_name,
                ref=f"A{header_row}:{end_col_letter}{data_end_row}"
            )
            table.tableStyleInfo = TableStyleInfo(
                name=table_style_name,
                showRowStripes=True,
                showColumnStripes=False
            )
            ws.add_table(table)

        ws.merge_cells("A1:R1")
        ws["A1"] = f"REKAPAN MENU RAMADAN - {selected_date}"
        ws["A1"].fill = title_fill
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24

        for table_key in ordered_selected_tables:
            table_config = DASHBOARD_TABLE_CONFIGS.get(table_key, {})
            column_defs = selected_column_defs.get(table_key) or table_config.get("columns", [])
            headers = [column["label"] for column in column_defs]
            table_rows = dashboard_table_rows.get(table_key, [])
            export_rows = [
                [row.get(column["key"], "-") for column in column_defs]
                for row in table_rows
            ]
            compact_number_cols = {
                index
                for index, column in enumerate(column_defs, start=1)
                if column.get("key") in table_config.get("compact_number_keys", set())
            }

            add_section_table(
                table_config.get("excel_title") or table_config.get("title") or table_key,
                headers,
                export_rows,
                table_config.get("excel_table_name") or f"Table{len(headers)}",
                table_config.get("excel_style_name") or "TableStyleMedium2",
                compact_number_cols=compact_number_cols
            )

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 22
        ws.column_dimensions["G"].width = 18
        ws.column_dimensions["H"].width = 22
        ws.column_dimensions["I"].width = 24
        ws.column_dimensions["J"].width = 14
        ws.column_dimensions["K"].width = 18
        ws.column_dimensions["L"].width = 22
        ws.column_dimensions["M"].width = 22
        ws.column_dimensions["N"].width = 22
        ws.column_dimensions["O"].width = 16
        ws.column_dimensions["P"].width = 18
        ws.column_dimensions["Q"].width = 28
        ws.column_dimensions["R"].width = 10

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        file_date = selected_date.replace("-", "")

        return send_file(
            output,
            download_name=f"rekapan_menu_ramadan_{file_date}.xlsx",
            as_attachment=True
        )
    #================== UPDATE STOCK =================
